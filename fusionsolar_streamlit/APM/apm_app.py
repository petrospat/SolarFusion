# apm_app.py
# FusionSolar Asset Performance Management Dashboard — Full APM Edition
# Tiers 1 + 2 + 3: KPI Scorecard, WCPR, Alerts, Revenue Waterfall,
#   Loss Cascade, Degradation, Capture Rate, Availability, PDF Report,
#   DSCR Monitor, ML Anomaly Detection, Production Forecast, OPEX Tracker,
#   Soiling Optimisation

import io
import os
import glob
import hmac
import time
import random
import calendar
import sqlite3
import urllib3
import warnings
import concurrent.futures
from datetime import datetime, timezone, date, timedelta
from typing import Tuple, List, Optional, Dict
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import requests
import streamlit as st
import plotly.graph_objects as go
import openpyxl
from plotly.subplots import make_subplots

warnings.filterwarnings("ignore")
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="FusionSolar APM",
    page_icon="☀️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# ACCESS GATE — this app has no other authentication layer of its own, and
# once deployed (e.g. Streamlit Community Cloud) its URL is reachable by
# anyone. Without this, any visitor could see real revenue/production
# figures, the plant's GPS location and legal entity name, trigger live
# FusionSolar/ENTSO-E API calls against the owner's own credentials, and
# write to the alerts/opex/revenue log with no attribution. A single shared
# password, checked before anything else renders, closes all of that off.
#
# Set in secrets.toml (locally) AND in the deployed app's Settings → Secrets
# (never commit a real value):
#     [app]
#     password = "choose-a-strong-password"
# ─────────────────────────────────────────────────────────────────────────────
def _check_app_password() -> bool:
    if st.session_state.get("_authenticated"):
        return True

    try:
        configured = st.secrets["app"]["password"]
    except Exception:
        configured = None

    if not configured:
        st.error(
            "⚠️ **No app password configured.** Add to `.streamlit/secrets.toml`:\n\n"
            "```toml\n[app]\npassword = \"choose-a-strong-password\"\n```\n\n"
            "For a deployed app, add the same under the app's **Settings → Secrets**."
        )
        return False

    st.markdown("### 🔒 FusionSolar APM — Sign in")
    with st.form("login_form"):
        entered = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Enter")
    if submitted:
        if hmac.compare_digest(entered, str(configured)):
            st.session_state["_authenticated"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False

if not _check_app_password():
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# PLANT CONSTANTS  ← configure these for your site
# ─────────────────────────────────────────────────────────────────────────────
PLANT_TZ          = "Europe/Athens"
PLANT_START_YEAR  = 2023          # COD year
PLANT_PEAK_KW     = 1000.0        # Installed DC capacity (kWp)
GAMMA             = -0.0035       # Temperature coefficient (%/°C) — typical mono-PERC
NOCT              = 45.0          # Nominal Operating Cell Temperature (°C)
INV_DEV_TYPE_IDS  = [1, 38, 39]   # Huawei device type IDs for inverters

# Finance constants (update per project)
ANNUAL_DEBT_SVC   = 120_000.0     # € — annual principal + interest
FIXED_OPEX        = 25_000.0      # € — fixed O&M, insurance, lease per year
VAR_OPEX_PER_MWH  = 2.5           # € per MWh variable O&M

# Alert thresholds
PR_ALERT_RED      = 0.65
PR_ALERT_AMBER    = 0.75
INV_TEMP_WARN     = 75.0
STRING_CV_WARN    = 0.12
PROD_VS_EXP_WARN  = 0.70          # flag if actual < 70% of PVGIS expected

# PVGIS-SARAH3 long-term monthly GTI (kWh/m²) — Asvestochori, Thessaloniki
PVGIS_GTI = {1:55.2,2:73.8,3:118.6,4:155.4,5:185.2,6:205.7,
             7:215.3,8:196.4,9:152.8,10:103.5,11:62.4,12:46.1}
T_AMB     = {1:4.5,2:5.8,3:9.3,4:14.6,5:20.1,6:25.2,
             7:27.8,8:27.4,9:22.5,10:16.2,11:10.6,12:6.1}

MONTH_LABELS = ["Jan","Feb","Mar","Apr","May","Jun",
                "Jul","Aug","Sep","Oct","Nov","Dec"]
PALETTE      = ["#f0b429","#3ecfcf","#60a5fa","#a78bfa","#fb923c","#34d399","#f472b6"]
ENTSOE_API   = "https://web-api.tp.entsoe.eu/api"   # SMP source
ENTSOE_ZONE  = "10YGR-HTSO-----Y"                   # Greece bidding zone EIC

# Shared, connection-pooled session for ENTSO-E requests. Reused across the
# parallel batch fetches so repeated calls don't each pay a fresh TLS handshake.
_ENTSOE_SESSION = requests.Session()
_ENTSOE_SESSION.mount("https://", requests.adapters.HTTPAdapter(pool_maxsize=10))


def _get_proxies() -> Optional[dict]:
    """Read proxy from secrets.toml [proxy] section or fall back to env vars."""
    try:
        p = st.secrets["proxy"]["https"]
        if p:
            if not p.startswith("http"):
                p = "http://" + p
            return {"https": p, "http": p}
    except Exception:
        pass
    return None

# Failure-predictor telemetry signals and their warning thresholds
FAILURE_SIGNALS = {
    "dataItemMap.temperature":  {"label":"Inverter Temp (°C)",   "warn":65.0,  "crit":75.0,  "unit":"°C",   "color":"#fb923c"},
    "dataItemMap.power_factor": {"label":"Power Factor",          "warn":0.90,  "crit":0.85,  "unit":"",     "color":"#60a5fa",  "low_bad":True},
    "dataItemMap.efficiency":   {"label":"Efficiency (%)",        "warn":95.0,  "crit":92.0,  "unit":"%",    "color":"#34d399",  "low_bad":True},
    "dataItemMap.elec_freq":    {"label":"Grid Frequency (Hz)",   "warn":49.8,  "crit":49.5,  "unit":"Hz",   "color":"#a78bfa"},
    "dataItemMap.active_power": {"label":"AC Power (kW)",         "warn":None,  "crit":None,  "unit":"kW",   "color":"#f0b429"},
}

# FusionSolar device-alarm severity codes (getAlarmList "lev" field).
# Huawei's Northbound API returns these as small ints; map to labels/colors
# for display. Unrecognised codes fall back to "Unknown" rather than erroring.
ALARM_LEVEL_MAP = {1:"Critical", 2:"Major", 3:"Minor", 4:"Warning"}
ALARM_LEVEL_COLOR = {"Critical":"#ff5f5f", "Major":"#fb923c", "Minor":"#f0b429",
                     "Warning":"#60a5fa", "Unknown":"#94a3b8"}
ALARM_LEVEL_ORDER = ["Critical","Major","Minor","Warning","Unknown"]

# ─────────────────────────────────────────────────────────────────────────────
# THEME HELPERS
# ─────────────────────────────────────────────────────────────────────────────
_BG  = "#0e1117"
_SRF = "#161b22"
_GRD = "#1f2333"
_TXT = "#e2e8f0"

def _base_layout(fig, title="", xaxis_title="", yaxis_title="",
                 height=None, barmode=None, showlegend=True):
    kw = dict(paper_bgcolor=_BG, plot_bgcolor=_BG,
              font=dict(color=_TXT),
              legend=dict(bgcolor="rgba(0,0,0,0)"),
              margin=dict(t=55,b=40,l=10,r=10),
              hovermode="x unified")
    if title:    kw["title"]   = dict(text=title, font=dict(size=15))
    if barmode:  kw["barmode"] = barmode
    if height:   kw["height"]  = height
    if not showlegend: kw["showlegend"] = False
    fig.update_layout(**kw)
    fig.update_xaxes(gridcolor=_GRD, zerolinecolor=_GRD,
                     title_text=xaxis_title, title_font=dict(color=_TXT))
    fig.update_yaxes(gridcolor=_GRD, zerolinecolor=_GRD,
                     title_text=yaxis_title, title_font=dict(color=_TXT))
    return fig

def _dual_layout(fig, title="", left_title="", right_title="",
                 left_color="#f0b429", right_color="#3ecfcf", height=None):
    kw = dict(paper_bgcolor=_BG, plot_bgcolor=_BG,
              font=dict(color=_TXT),
              legend=dict(bgcolor="rgba(0,0,0,0)", orientation="h",
                          yanchor="bottom", y=1.02),
              margin=dict(t=80,b=40,l=70,r=70),
              hovermode="x unified")
    if title:  kw["title"]  = dict(text=title, font=dict(size=15))
    if height: kw["height"] = height
    fig.update_layout(**kw)
    fig.update_xaxes(gridcolor=_GRD, zerolinecolor=_GRD, tickformat="%H:%M")
    fig.update_yaxes(title_text=left_title, gridcolor=_GRD, zerolinecolor=_GRD,
                     rangemode="tozero",
                     title_font=dict(color=left_color),
                     tickfont=dict(color=left_color), secondary_y=False)
    fig.update_yaxes(title_text=right_title, gridcolor="rgba(0,0,0,0)",
                     zerolinecolor=_GRD,
                     rangemode="tozero",
                     title_font=dict(color=right_color),
                     tickfont=dict(color=right_color), secondary_y=True)
    return fig

# ─────────────────────────────────────────────────────────────────────────────
# SQLITE  — alerts, opex, downtime, cleaning events
# ─────────────────────────────────────────────────────────────────────────────
# Resolved relative to this file, not the process's working directory: a bare
# "apm_data.db" would silently create/read a different, empty database if the
# app is ever launched from another cwd (exactly what Streamlit Community
# Cloud does for a nested main file path like fusionsolar_streamlit/APM/
# apm_app.py — it runs with cwd at the repo root) — caught live when a stray
# near-empty apm_data.db at the repo root shadowed the real, fully-backfilled
# one here after adding the password gate prompted a fresh run.
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "apm_data.db")

# Locally exported per-inverter Huawei Excel files (com1-1/2/3), genuinely
# native 15-min resolution. Preferred over the live FusionSolar API for the
# Intraday tab wherever a day is covered: getKpiStation5min doesn't exist
# for this account (confirmed HTTP 404, not a rate limit — so the live API
# path can never return better than hourly), and getKpiStationHour itself is
# frequently rate-limited (failCode=407) on top of that.
LOCAL_PROD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                 "..", "Historical prod data 25-26")


@st.cache_data(show_spinner=False)
def _load_local_inverter_production() -> pd.DataFrame:
    """Sum of com1-1/com1-2/com1-3 Active power(kW) at every 15-min
    timestamp found in the local exports. Returns DataFrame[dt, kw_sum]
    (dt tz-aware, Europe/Athens) — empty if no files are present.

    File quirk: these exports don't declare a valid <dimension> in their
    sheet XML, so openpyxl's read_only max_row inference reports 1 unless
    max_row is forced higher than the true data (iter_rows below).

    Start-Time convention: "YYYY-MM-DD HH:MM:SS DST" during EEST months,
    plain "YYYY-MM-DD HH:MM:SS" (EET) otherwise — Huawei's own resolved
    local time, used directly via tz_localize(ambiguous=...) rather than
    re-deriving DST from the naive timestamp.
    """
    files = sorted(glob.glob(os.path.join(LOCAL_PROD_FOLDER, "Inverter_*.xlsx")))
    if not files:
        return pd.DataFrame(columns=["dt", "kw_sum"])
    ts_list, dst_list, kw_list = [], [], []
    for fp in files:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=5, max_row=200000, max_col=8, values_only=True):
            start_time, power = row[3], row[4]
            if not start_time or power is None:
                continue
            is_dst = start_time.endswith(" DST")
            ts_list.append(start_time[:-4] if is_dst else start_time)
            dst_list.append(is_dst)
            kw_list.append(float(power))
        wb.close()
    df = pd.DataFrame({"ts_str": ts_list, "is_dst": dst_list, "kw": kw_list})
    naive = pd.to_datetime(df["ts_str"], format="%Y-%m-%d %H:%M:%S")
    df["dt"] = naive.dt.tz_localize(PLANT_TZ, ambiguous=df["is_dst"].to_numpy(),
                                    nonexistent="shift_forward")
    return (df.groupby("dt")["kw"].sum().reset_index()
           .rename(columns={"kw": "kw_sum"}).sort_values("dt").reset_index(drop=True))


def _local_prod_json_for_day(target_date: date) -> Optional[dict]:
    """Builds the same {"data": [...], "_src": ...} shape api_15min() returns,
    from local Excel data, so the rest of the Intraday tab (and
    _compute_day_revenue_from_frames) doesn't need to know the source."""
    site = _load_local_inverter_production()
    if site.empty:
        return None
    day_df = site[site["dt"].dt.date == target_date]
    if day_df.empty:
        return None
    rows = [{"collectTime": int(dt.tz_convert("UTC").timestamp() * 1000),
            "active_power": float(kw)} for dt, kw in zip(day_df["dt"], day_df["kw_sum"])]
    return {"data": rows, "_src": "15min"}

def _get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("""CREATE TABLE IF NOT EXISTS alerts(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ts TEXT, severity TEXT, category TEXT,
        message TEXT, status TEXT DEFAULT 'Open')""")
    conn.execute("""CREATE TABLE IF NOT EXISTS opex(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        dt TEXT, category TEXT, description TEXT, amount_eur REAL)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS downtime(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        start_dt TEXT, end_dt TEXT, inverter TEXT,
        lost_kwh REAL, cause TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS cleaning(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        dt TEXT, cost_eur REAL, yield_recovery_pct REAL, notes TEXT)""")
    # Telemetry history: snapshots of inverter health signals over time
    conn.execute("""CREATE TABLE IF NOT EXISTS telemetry_history(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ts TEXT, inverter_id TEXT, signal TEXT, value REAL)""")
    # Daily revenue log: stores per-day revenue from intraday 15-min calculation
    conn.execute("""CREATE TABLE IF NOT EXISTS daily_revenue(
        day TEXT PRIMARY KEY,
        kwh REAL, revenue_eur REAL, avg_price REAL, fetched_ts TEXT)""")
    # 15-min revenue timeseries: one row per 15-min bucket — generation,
    # ENTSO-E price, and revenue for that interval. This is the granular
    # source daily_revenue is aggregated from.
    conn.execute("""CREATE TABLE IF NOT EXISTS revenue_15min(
        dt TEXT PRIMARY KEY,
        day TEXT, kwh REAL, price_eur_mwh REAL, revenue_eur REAL, fetched_ts TEXT)""")
    conn.execute("""CREATE INDEX IF NOT EXISTS idx_revenue_15min_day
        ON revenue_15min(day)""")
    # Price cache: persists last-fetched DAM prices for offline fallback
    conn.execute("""CREATE TABLE IF NOT EXISTS price_cache(
        ym TEXT PRIMARY KEY, avg_price_eur_mwh REAL, fetched_ts TEXT)""")
    # Durable DAM price journal: every 15-min/hourly price ever successfully
    # fetched from ENTSO-E, kept permanently (not just this session/day). When
    # a live ENTSO-E pull fails, this is what lets the app serve the last
    # known-good value for that exact timestamp instead of erroring out —
    # there is no free public secondary live source for Greek DAM prices
    # (IPTO only publishes grid-operation data; the exchange's own API is a
    # certified-trading-member-only order-execution interface), so a durable
    # local cache is the realistic fallback rather than a second live API.
    conn.execute("""CREATE TABLE IF NOT EXISTS dam_prices(
        timestamp_utc TEXT NOT NULL, bidding_zone TEXT NOT NULL,
        price_eur_mwh REAL, resolution_min INTEGER,
        source TEXT NOT NULL, fetched_ts TEXT NOT NULL,
        PRIMARY KEY (timestamp_utc, bidding_zone))""")
    conn.commit()
    return conn


def _save_telemetry_snapshot(df_kpi: pd.DataFrame):
    """Persist current inverter telemetry to history table for trend analysis."""
    if df_kpi.empty: return
    conn = _get_db()
    ts = datetime.now().isoformat()
    rows = []
    inv_id_col = next((c for c in df_kpi.columns
                       if "id" in c.lower() or "sn" in c.lower() or "dev" in c.lower()), None)
    for idx, row in df_kpi.iterrows():
        inv_id = str(row[inv_id_col]) if inv_id_col else str(idx)
        for sig_col, meta in FAILURE_SIGNALS.items():
            if sig_col in df_kpi.columns:
                val = pd.to_numeric(row.get(sig_col), errors="coerce")
                if pd.notna(val):
                    rows.append((ts, inv_id, meta["label"], float(val)))
    if rows:
        conn.executemany(
            "INSERT INTO telemetry_history(ts,inverter_id,signal,value) VALUES(?,?,?,?)",
            rows)
        conn.commit()
    conn.close()


def _cache_dam_price(ym: str, price: float):
    conn = _get_db()
    conn.execute("INSERT OR REPLACE INTO price_cache VALUES(?,?,?)",
                 (ym, price, datetime.now().isoformat()))
    conn.commit(); conn.close()


def _get_cached_dam_price(ym: str) -> Optional[float]:
    conn = _get_db()
    row = conn.execute("SELECT avg_price_eur_mwh FROM price_cache WHERE ym=?",
                       (ym,)).fetchone()
    conn.close()
    return row[0] if row else None


def _persist_dam_prices(df: pd.DataFrame, resolution_min: int, source: str = "ENTSOE") -> None:
    """
    Durably journal a batch of [dt, price] rows (any tz — converted to UTC
    here) into dam_prices. No Streamlit calls — safe from worker threads.
    ON CONFLICT policy mirrors the source-priority pattern used elsewhere in
    this app: 'ENTSOE' always overwrites (it's the live/authoritative pull);
    anything else only fills a slot that's genuinely empty, so a fallback
    read never silently degrades an already-good value.
    """
    if df.empty:
        return
    fetched_ts = datetime.now(timezone.utc).isoformat()
    rows = [
        (pd.Timestamp(r.dt).tz_convert("UTC").isoformat(), ENTSOE_ZONE,
         float(r.price), resolution_min, source, fetched_ts)
        for r in df.itertuples(index=False) if pd.notna(r.price)
    ]
    if not rows:
        return
    conn = _get_db()
    conn.executemany("""
        INSERT INTO dam_prices
            (timestamp_utc, bidding_zone, price_eur_mwh, resolution_min, source, fetched_ts)
        VALUES (?,?,?,?,?,?)
        ON CONFLICT(timestamp_utc, bidding_zone) DO UPDATE SET
            price_eur_mwh  = CASE WHEN excluded.source='ENTSOE' THEN excluded.price_eur_mwh
                                  WHEN dam_prices.price_eur_mwh IS NULL THEN excluded.price_eur_mwh
                                  ELSE dam_prices.price_eur_mwh END,
            resolution_min = CASE WHEN excluded.source='ENTSOE' THEN excluded.resolution_min
                                  WHEN dam_prices.price_eur_mwh IS NULL THEN excluded.resolution_min
                                  ELSE dam_prices.resolution_min END,
            source         = CASE WHEN excluded.source='ENTSOE' THEN excluded.source
                                  WHEN dam_prices.price_eur_mwh IS NULL THEN excluded.source
                                  ELSE dam_prices.source END,
            fetched_ts     = CASE WHEN excluded.source='ENTSOE' THEN excluded.fetched_ts
                                  WHEN dam_prices.price_eur_mwh IS NULL THEN excluded.fetched_ts
                                  ELSE dam_prices.fetched_ts END
    """, rows)
    conn.commit(); conn.close()


def _read_cached_dam_prices(target_date: date) -> pd.DataFrame:
    """
    Read back a day's durably-journaled DAM prices (whatever source last
    filled each slot), shaped to match _fetch_dam_daily_uncached's normal
    return: DataFrame[dt, price, _src] in PLANT_TZ. Empty if nothing was
    ever successfully cached for this day.
    """
    day_start = datetime(target_date.year, target_date.month, target_date.day, tzinfo=timezone.utc)
    day_end = day_start + timedelta(days=1)
    conn = _get_db()
    df = pd.read_sql_query(
        """SELECT timestamp_utc, price_eur_mwh, source FROM dam_prices
           WHERE bidding_zone=? AND timestamp_utc>=? AND timestamp_utc<?
           ORDER BY timestamp_utc""",
        conn, params=(ENTSOE_ZONE, day_start.isoformat(), day_end.isoformat()))
    conn.close()
    if df.empty:
        return pd.DataFrame(columns=["dt", "price", "_src"])
    df["dt"]    = pd.to_datetime(df["timestamp_utc"], utc=True).dt.tz_convert(PLANT_TZ)
    df["price"] = df["price_eur_mwh"]
    df["_src"]  = "local-cache:" + df["source"]
    return df[["dt", "price", "_src"]].sort_values("dt").reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# ENTSO-E CONNECTIVITY PROBE
# ─────────────────────────────────────────────────────────────────────────────
def probe_entsoe(timeout: int = 10) -> dict:
    """Test ENTSO-E API connectivity and return a diagnostic result dict."""
    result = dict(reachable=False, status_code=None, error_type=None,
                  user_message=None, price_parseable=False,
                  sample_price=None, n_periods=0)

    token = _get_entsoe_token()
    if not token:
        result["error_type"] = "NO_TOKEN"
        result["user_message"] = (
            "⚠️ **No ENTSO-E API token configured.**\n\n"
            "Add to `.streamlit/secrets.toml`:\n"
            "```toml\n[entsoe]\napi_key = \"your-token-here\"\n```\n"
            "Register free at [transparency.entsoe.eu](https://transparency.entsoe.eu)")
        return result

    test_date    = date.today() - timedelta(days=1)
    period_start = datetime(test_date.year, test_date.month,
                            test_date.day, 0, 0, tzinfo=timezone.utc)
    period_end   = period_start + timedelta(days=1)

    # Greece Day-Ahead prices may use A01 or A16 — try both
    for process_type in ["A01", "A16"]:
        params = {
            "securityToken": token,
            "documentType":  "A44",
            "processType":   process_type,
            "in_Domain":     ENTSOE_ZONE,
            "out_Domain":    ENTSOE_ZONE,
            "periodStart":   period_start.strftime("%Y%m%d%H%M"),
            "periodEnd":     period_end.strftime("%Y%m%d%H%M"),
        }
        try:
            r = requests.get(ENTSOE_API, params=params, timeout=timeout,
                             proxies=_get_proxies(), verify=False)
            result["status_code"] = r.status_code
            result["reachable"]   = True

            if r.status_code == 401:
                result["error_type"]   = "INVALID_TOKEN"
                result["user_message"] = (
                    "❌ **401 Unauthorised** — token invalid or expired. "
                    "Regenerate at transparency.entsoe.eu and update secrets.toml.")
                return result

            if r.status_code != 200:
                result["error_type"]   = "HTTP_ERROR"
                result["user_message"] = f"⚠️ ENTSO-E returned HTTP {r.status_code}"
                continue

            import xml.etree.ElementTree as ET
            root     = ET.fromstring(r.content)
            root_tag = root.tag
            actual_ns = root_tag.split("}")[0].lstrip("{") if "}" in root_tag else ""

            result["raw_xml"]     = r.text[:800]
            result["detected_ns"] = actual_ns or "(none)"
            result["root_tag"]    = root_tag

            if "acknowledgement" in actual_ns.lower():
                ns_ack = {"a": actual_ns}
                reason = (root.findtext(".//a:Reason/a:text", namespaces=ns_ack)
                          or root.findtext(".//a:reason", namespaces=ns_ack) or "")
                code   = (root.findtext(".//a:Reason/a:code", namespaces=ns_ack)
                          or root.findtext(".//a:code", namespaces=ns_ack) or "")
                result["ack_reason"] = reason
                result["ack_code"]   = code
                # code 999 = no data, try next process type
                continue

            ns     = {"ns": actual_ns} if actual_ns else {}
            prefix = "ns:" if actual_ns else ""
            prices = [float(pt.findtext(f"{prefix}price.amount", namespaces=ns))
                      for pt in root.findall(f".//{prefix}Point", ns)
                      if pt.findtext(f"{prefix}price.amount", namespaces=ns)]

            if prices:
                result["price_parseable"]  = True
                result["sample_price"]     = float(np.mean(prices))
                result["n_periods"]        = len(prices)
                result["working_process"]  = process_type
                result["error_type"]       = None
                result["user_message"]     = (
                    f"✅ **ENTSO-E API operational** — Greece SMP (A44/{process_type}).\n"
                    f"Date: `{test_date}` — **{len(prices)} periods**, "
                    f"avg **{result['sample_price']:.2f} €/MWh**.")
                return result

        except Exception as e:
            msg = _redact_token(str(e), token)
            result["error_type"]   = "EXCEPTION"
            result["user_message"] = f"❌ `{msg[:200]}`"
            continue

    # All process types exhausted
    if not result.get("price_parseable"):
        result["error_type"]   = result.get("error_type") or "NO_PRICES"
        result["user_message"] = (
            f"⚠️ Connected (HTTP 200) but no prices found for `{test_date}` "
            f"with processType A01 or A16.\n\n"
            f"**ACK code:** `{result.get('ack_code','?')}` — "
            f"**Reason:** {result.get('ack_reason','No data returned.')}\n\n"
            f"Try opening the Raw XML expander below to inspect the full response.")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# API BACKOFF
# ─────────────────────────────────────────────────────────────────────────────
def _post(session, url, payload, timeout=25, retries=3):
    last = None
    for i in range(retries):
        try:
            r = session.post(url, json=payload, timeout=timeout)
            last = r
            if r.status_code == 200:
                j = r.json()
                if j.get("failCode") == 407:
                    time.sleep(min(30, 2*(2**i)) + random.uniform(0,1))
                    continue
                return j
        except Exception:
            pass
        time.sleep(min(30, 2*(2**i)) + random.uniform(0,1))
    try:    return last.json()
    except: return {"data": None, "failCode": -1}

# ─────────────────────────────────────────────────────────────────────────────
# HUAWEI CLIENT
# ─────────────────────────────────────────────────────────────────────────────
class HuaweiClient:
    def __init__(self, cfg):
        # Accept both naming styles: host/user/password OR base_url/username/system_code
        self.base_url    = (cfg.get("base_url") or cfg.get("host", "")).rstrip("/")
        self.username    = cfg.get("username") or cfg.get("user", "")
        self.system_code = cfg.get("system_code") or cfg.get("password", "")
        # Default verify_ssl to False — FusionSolar EU hosts often fail local CA checks
        self.verify_ssl  = cfg.get("verify_ssl", False)
        self.s = requests.Session()
        self.s.verify = self.verify_ssl
        self.s.headers.update({"Content-Type":"application/json",
                               "User-Agent":"Streamlit-APM"})

    def login(self) -> Tuple[bool, str]:
        try:
            r = self.s.post(f"{self.base_url}/thirdData/login",
                            json={"userName":self.username,
                                  "systemCode":self.system_code}, timeout=15)
            tok = r.cookies.get("XSRF-TOKEN") or r.headers.get("xsrf-token")
            if not tok: return False, "No XSRF token — check credentials."
            self.s.headers.update({"XSRF-TOKEN": tok})
            return True, "OK"
        except Exception as e:
            return False, str(e)

    def get_stations(self):
        r = _post(self.s, f"{self.base_url}/thirdData/getStationList", {})
        d = r.get("data", [])
        rows = d.get("list",[]) if isinstance(d,dict) else (d if isinstance(d,list) else [])
        return (rows, None) if rows else (None, "No stations returned")

    @property
    def xsrf(self): return self.s.headers.get("XSRF-TOKEN","")

def ensure_client():
    if "hw_client" not in st.session_state:
        try:    cfg = st.secrets["fusion"]
        except: st.error("❌ `[fusion]` missing from secrets.toml"); return None, None
        c = HuaweiClient(cfg)
        ok, msg = c.login()
        if not ok: st.error(f"❌ Login: {msg}"); return None, None
        stations, err = c.get_stations()
        if not stations: st.error(f"❌ Stations: {err}"); return None, None
        st.session_state["hw_client"]   = c
        st.session_state["hw_stations"] = stations
    return st.session_state["hw_client"], st.session_state["hw_stations"]

# ─────────────────────────────────────────────────────────────────────────────
# DATA HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _norm(rows):
    if not rows or not isinstance(rows, (list, dict)):
        # A malformed/error response (e.g. a string message instead of a
        # record list) would otherwise make pd.json_normalize raise
        # NotImplementedError — degrade to empty instead of crashing the app.
        return pd.DataFrame()
    df = pd.json_normalize(rows, sep=".")
    return df.loc[:, ~df.columns.duplicated()].copy()

def _resolve(df):
    t = next((c for c in df.columns if "time" in c.lower() or "collect" in c.lower()), None)
    e = next((c for c in df.columns if "inverterYield" in c or "month_cap" in c
              or "energy" in c.lower()), None)
    return t, e

def get_budget(year):
    tgts = [117077,89742,140573,172775,177950,186287,
            197265,190014,168524,132649,86079,82732]
    if year == 2025:
        for i in range(9): tgts[i] = 0
    return pd.DataFrame({"Month":MONTH_LABELS,"Budget_kWh":tgts,
                          "MonthNum":range(1,13),"Year":year})

def pvgis_df(years, ref_pr):
    rows=[]
    for yr in years:
        for m in range(1,13):
            g = PVGIS_GTI[m]
            rows.append({"YearMonth":f"{yr}-{m:02d}","MonthNum":m,"Year":yr,
                         "GTI":g,"T_amb":T_AMB[m],
                         "Expected_kWh": g * PLANT_PEAK_KW * ref_pr})
    return pd.DataFrame(rows)

def wcpr(energy_kwh, gti_kwh_m2, t_amb_c):
    """
    Weather-corrected PR — IEC 61724.
    T_cell estimated from T_amb + NOCT proxy.
    """
    t_cell = t_amb_c + (NOCT - 20) / 800 * (gti_kwh_m2 * 1000 / 730)  # rough irr proxy
    ref = gti_kwh_m2 * PLANT_PEAK_KW * (1 + GAMMA * (t_cell - 25))
    return energy_kwh / ref if ref > 0 else np.nan

def add_alert(severity, category, message):
    conn = _get_db()
    conn.execute("INSERT INTO alerts(ts,severity,category,message) VALUES(?,?,?,?)",
                 (datetime.now().isoformat(), severity, category, message))
    conn.commit()
    conn.close()

# ─────────────────────────────────────────────────────────────────────────────
# CACHED API CALLS
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=3600, show_spinner=False)
def api_monthly(base_url, sid, year, xsrf, verify):
    s = requests.Session(); s.verify=verify
    s.headers.update({"Content-Type":"application/json","XSRF-TOKEN":xsrf})
    dt = datetime(year,1,1,tzinfo=timezone.utc)
    j  = _post(s, f"{base_url}/thirdData/getKpiStationMonth",
               {"stationCodes":sid,"collectTime":int(dt.timestamp()*1000)})
    raw = j.get("data",[])
    if raw and isinstance(raw,list) and "kpiList" in raw[0]: raw=raw[0]["kpiList"]
    return _norm(raw)

def api_monthly_years(base_url, sid, years, xsrf, verify) -> Dict[int, pd.DataFrame]:
    """Fetch api_monthly() for several years concurrently instead of one-by-one."""
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(4, len(years))) as ex:
        fut_map = {ex.submit(api_monthly, base_url, sid, yr, xsrf, verify): yr
                   for yr in years}
        return {fut_map[fut]: fut.result() for fut in concurrent.futures.as_completed(fut_map)}

@st.cache_data(ttl=1800, show_spinner=False)
def api_dev_5min(base_url, sid, target_date, xsrf, verify):
    """
    Device-level 5-min production (thirdData/getDevFiveMinutes), summed
    across all inverters. Confirmed via live testing to be genuinely native
    5-min resolution AND — unlike the station-level getKpiStation5min
    (HTTP 404, not provisioned for this account) / getKpiStationHour
    (frequently failCode=407) — not subject to the same account-wide rate
    limit: 10 consecutive calls across different days, no delay, all
    succeeded. Returns the same {"data": [...], "_src": ...} shape
    api_15min() does, so it's a drop-in primary source for it.
    """
    s = requests.Session(); s.verify=verify
    s.headers.update({"Content-Type":"application/json","XSRF-TOKEN":xsrf})
    j_devs = _post(s, f"{base_url}/thirdData/getDevList", {"stationCodes":sid})
    devs = j_devs.get("data") or []
    inv_ids = [d["id"] for d in devs if d.get("devTypeId") in INV_DEV_TYPE_IDS]
    if not inv_ids:
        return {}
    ms = int(datetime.combine(target_date, datetime.min.time(),
                              tzinfo=timezone.utc).timestamp()*1000)
    j = _post(s, f"{base_url}/thirdData/getDevFiveMinutes",
             {"devIds":",".join(map(str,inv_ids)),"devTypeId":1,"collectTime":ms})
    raw = j.get("data")
    if not raw or not isinstance(raw, list):
        return {}
    df = pd.json_normalize(raw, sep=".")
    # Exact segment match, not endswith: "reactive_power" also ends with the
    # substring "active_power" and — caught live — sorts before it in this
    # payload's column order, so an endswith() match silently summed
    # reactive power (which nets negative) instead of active power.
    pc = next((c for c in df.columns if c.split(".")[-1] == "active_power"), None)
    if not pc or "collectTime" not in df.columns:
        return {}
    df[pc] = pd.to_numeric(df[pc], errors="coerce")
    # Sum the 3 inverters' power per collectTime into a single site-level
    # reading — must happen before this goes anywhere near
    # _compute_day_revenue_from_frames, since its median-gap interval
    # detection assumes one row per timestamp (3 duplicate-timestamp rows
    # per 5-min tick would corrupt that into a near-zero median).
    site = (df.groupby("collectTime")[pc].sum().reset_index()
           .rename(columns={pc:"active_power"}))
    return {"data": site.to_dict("records"), "_src": "dev5min"}


@st.cache_data(ttl=1800, show_spinner=False)
def api_15min(base_url, sid, target_date, xsrf, verify):
    jd = api_dev_5min(base_url, sid, target_date, xsrf, verify)
    if jd.get("data"):
        return jd
    s = requests.Session(); s.verify=verify
    s.headers.update({"Content-Type":"application/json","XSRF-TOKEN":xsrf})
    ms = int(datetime.combine(target_date, datetime.min.time(),
                              tzinfo=timezone.utc).timestamp()*1000)
    j5 = _post(s, f"{base_url}/thirdData/getKpiStation5min",
               {"stationCodes":sid,"collectTime":ms})
    if j5.get("data"): j5["_src"]="5min"; return j5
    jh = _post(s, f"{base_url}/thirdData/getKpiStationHour",
               {"stationCodes":sid,"collectTime":ms})
    jh["_src"]="hour"; return jh

@st.cache_data(ttl=300, show_spinner=False)
def api_realtime(base_url, dev_type, dev_ids, xsrf, verify):
    s = requests.Session(); s.verify=verify
    s.headers.update({"Content-Type":"application/json","XSRF-TOKEN":xsrf})
    j = _post(s, f"{base_url}/thirdData/getDevRealKpi",
              {"devTypeId":dev_type,"devIds":",".join(map(str,dev_ids))})
    return _norm(j.get("data",[]))

# ── Device alarms / events ──────────────────────────────────────────────────
# NOTE: Huawei's Northbound "thirdData/getAlarmList" field names vary a bit
# by tenant/API version, so columns are resolved by keyword (_resolve_alarm_cols)
# rather than hardcoded — the Events tab's "Raw response" expander lets you
# confirm/adjust the mapping against your actual account's payload.
def _resolve_alarm_cols(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    def find(*keys):
        for c in df.columns:
            cl = c.lower()
            if all(k in cl for k in keys):
                return c
        return None
    return dict(
        time   = find("raise","time") or find("occur","time") or find("time"),
        name   = find("alarm","name") or find("name"),
        level  = find("lev") or find("severity"),
        device = find("device","name") or find("station","name") or find("device"),
        cause  = find("cause") or find("desc") or find("suggestion"),
        status = find("status") or find("clear"),
    )

@st.cache_data(ttl=600, show_spinner=False)
def api_alarms(base_url, sid, begin_ms, end_ms, xsrf, verify):
    """
    Fetch device alarms/events for a station over [begin_ms, end_ms) from
    FusionSolar (thirdData/getAlarmList). Returns a raw normalised DataFrame —
    pass through normalize_alarms() to get the stable [dt, Device, Severity,
    Alarm, Cause, Status] schema used by the UI.
    """
    s = requests.Session(); s.verify=verify
    s.headers.update({"Content-Type":"application/json","XSRF-TOKEN":xsrf})
    j = _post(s, f"{base_url}/thirdData/getAlarmList",
              {"stationCodes":sid, "beginTime":begin_ms, "endTime":end_ms,
               "language":"en_US"})
    # Some Huawei endpoints (e.g. getStationList) wrap results as
    # {"data": {"list": [...], "total": N, ...}} rather than a bare list —
    # getAlarmList appears to follow the same pagination pattern. Unwrap it,
    # and don't assume the shape: anything else degrades to no rows instead
    # of crashing (see _norm's defensive check for the same reason).
    d = j.get("data", [])
    rows = d.get("list", []) if isinstance(d, dict) else (d if isinstance(d, list) else [])
    return _norm(rows)

def api_alarms_range(base_url, sid, start_date, end_date, xsrf, verify,
                     chunk_days=30) -> pd.DataFrame:
    """
    Fetch alarms over [start_date, end_date] (inclusive), chunked into
    <=chunk_days windows and fetched concurrently — mirrors the ENTSO-E/KPI
    batch helpers above so wide date ranges don't serialize into a long wait.
    """
    windows = []
    d = start_date
    while d <= end_date:
        w_end = min(d + timedelta(days=chunk_days-1), end_date)
        windows.append((d, w_end))
        d = w_end + timedelta(days=1)

    def _fetch(w):
        cs, ce = w
        b = int(datetime.combine(cs, datetime.min.time(), tzinfo=timezone.utc).timestamp()*1000)
        e = int(datetime.combine(ce + timedelta(days=1), datetime.min.time(),
                                 tzinfo=timezone.utc).timestamp()*1000) - 1
        return api_alarms(base_url, sid, b, e, xsrf, verify)

    if len(windows) == 1:
        return _fetch(windows[0])

    with concurrent.futures.ThreadPoolExecutor(max_workers=min(6, len(windows))) as ex:
        dfs = list(ex.map(_fetch, windows))
    dfs = [d for d in dfs if not d.empty]
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

def normalize_alarms(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Map a raw getAlarmList DataFrame onto a stable schema:
    [dt, Device, Severity, Alarm, Cause, Status] — used by both the Intraday
    hover overlay and the Events Analytics tab.
    """
    empty = pd.DataFrame(columns=["dt","Device","Severity","Alarm","Cause","Status"])
    if df_raw.empty:
        return empty
    cols = _resolve_alarm_cols(df_raw)
    if not cols["time"] or not cols["name"]:
        return empty

    out = pd.DataFrame()
    ts = pd.to_numeric(df_raw[cols["time"]], errors="coerce")
    out["dt"] = pd.to_datetime(ts, unit="ms", utc=True, errors="coerce").dt.tz_convert(PLANT_TZ)
    out["Device"] = df_raw[cols["device"]].astype(str) if cols["device"] else "—"
    if cols["level"]:
        lv = pd.to_numeric(df_raw[cols["level"]], errors="coerce")
        out["Severity"] = lv.map(ALARM_LEVEL_MAP).fillna("Unknown")
    else:
        out["Severity"] = "Unknown"
    out["Alarm"] = df_raw[cols["name"]].astype(str)
    out["Cause"] = df_raw[cols["cause"]].astype(str) if cols["cause"] else ""
    out["Status"] = df_raw[cols["status"]].astype(str) if cols["status"] else ""
    return out.dropna(subset=["dt"]).sort_values("dt").reset_index(drop=True)

def _add_event_markers(fig, events: pd.DataFrame, secondary_y=False, y_frac=0.04):
    """
    Overlay events as a hoverable scatter "rug" near the bottom of a
    time-series figure — one trace per severity so the legend can toggle
    them and the marker color communicates severity at a glance.
    """
    if events.empty:
        return fig
    ymax = 0.0
    for tr in fig.data:
        if getattr(tr, "y", None) is not None and len(tr.y):
            vals = pd.to_numeric(pd.Series(tr.y), errors="coerce").dropna()
            if not vals.empty:
                ymax = max(ymax, float(vals.max()))
    y_pos = ymax * y_frac if ymax > 0 else 0
    for sev in ALARM_LEVEL_ORDER:
        sub = events[events["Severity"] == sev]
        if sub.empty:
            continue
        fig.add_scatter(
            x=sub["dt"], y=[y_pos]*len(sub), mode="markers",
            name=f"⚠ {sev}",
            marker=dict(symbol="diamond", size=9,
                       color=ALARM_LEVEL_COLOR.get(sev, "#94a3b8"),
                       line=dict(width=1, color="#0e1117")),
            customdata=sub[["Device","Alarm","Cause"]].values,
            hovertemplate=(f"<b>{sev}</b> — %{{customdata[1]}}<br>"
                          "%{customdata[0]}<br>%{customdata[2]}<br>"
                          "%{x}<extra></extra>"),
            secondary_y=secondary_y)
    return fig

# ── Price source helpers ──────────────────────────────────────────────────────
# Greek SMP source: ENTSO-E Transparency Platform, document A44 (Day-Ahead Prices)
# Domain: 10YGR-HTSO-----Y  (Greece bidding zone)
# Token : secrets.toml → [entsoe] api_key
# Register free at: https://transparency.entsoe.eu

def _get_entsoe_token() -> Optional[str]:
    """Read ENTSO-E API token from Streamlit secrets."""
    try:
        return st.secrets["entsoe"]["api_key"]
    except Exception:
        return None


def _redact_token(msg: str, token: Optional[str]) -> str:
    """
    Strip the ENTSO-E securityToken out of exception text before it's stored
    in session_state or shown in the UI. requests/urllib3 connection errors
    (proxy failures, timeouts, etc.) embed the full request URL — including
    the query string — in their message, so the token leaks unless removed.
    """
    if token and token in msg:
        msg = msg.replace(token, "***")
    return msg


def _fetch_dam_daily_uncached(target_date: date, token: str,
                              timeout: int = 20) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Raw ENTSO-E 15-min SMP fetch for one day — no Streamlit calls inside, so
    it's safe to run from a worker thread (used by dam_daily / dam_daily_batch).
    Returns (DataFrame[dt, price, _src], error_message_or_None).
    """
    empty = pd.DataFrame(columns=["dt", "price", "_src"])
    period_start = datetime(target_date.year, target_date.month,
                            target_date.day, 0, 0, tzinfo=timezone.utc)
    period_end   = period_start + timedelta(days=1)

    import xml.etree.ElementTree as ET

    last_err = None
    for process_type in ["A01", "A16"]:
        params = {
            "securityToken": token,
            "documentType":  "A44",
            "processType":   process_type,
            "in_Domain":     ENTSOE_ZONE,
            "out_Domain":    ENTSOE_ZONE,
            "periodStart":   period_start.strftime("%Y%m%d%H%M"),
            "periodEnd":     period_end.strftime("%Y%m%d%H%M"),
        }
        try:
            r = _ENTSOE_SESSION.get(ENTSOE_API, params=params, timeout=timeout,
                                    proxies=_get_proxies(), verify=False)
            if r.status_code == 401:
                return empty, "ENTSO-E 401 Unauthorised — check api_key in secrets.toml"
            if r.status_code != 200:
                last_err = f"ENTSO-E HTTP {r.status_code}"
                continue

            root     = ET.fromstring(r.content)
            root_tag = root.tag
            actual_ns = root_tag.split("}")[0].lstrip("{") if "}" in root_tag else ""

            if "acknowledgement" in actual_ns.lower():
                continue   # no data for this process type, try next

            ns     = {"ns": actual_ns} if actual_ns else {}
            prefix = "ns:" if actual_ns else ""

            rows = []
            for ts in root.findall(f".//{prefix}TimeSeries", ns):
                resolution = ts.findtext(f".//{prefix}resolution", namespaces=ns) or "PT60M"
                freq_min   = 15 if "PT15M" in resolution else 60
                start_str  = ts.findtext(f".//{prefix}timeInterval/{prefix}start", namespaces=ns)
                if not start_str:
                    continue
                ts_start = datetime.strptime(start_str, "%Y-%m-%dT%H:%MZ").replace(
                    tzinfo=timezone.utc)
                for pt in ts.findall(f".//{prefix}Point", ns):
                    pos   = int(pt.findtext(f"{prefix}position", namespaces=ns) or 0)
                    price = float(pt.findtext(f"{prefix}price.amount", namespaces=ns) or "nan")
                    dt    = ts_start + timedelta(minutes=freq_min * (pos - 1))
                    rows.append({"dt": dt, "price": price})

            if rows:
                df = pd.DataFrame(rows)
                df["dt"] = pd.to_datetime(df["dt"]).dt.tz_convert(PLANT_TZ)
                df = df.dropna(subset=["price"]).sort_values("dt").reset_index(drop=True)
                df["_src"] = f"ENTSO-E:A44/{process_type}"
                _persist_dam_prices(df[["dt", "price"]],
                                   resolution_min=15 if freq_min == 15 else 60)
                return df[["dt", "price", "_src"]], None

        except Exception as e:
            safe_msg = _redact_token(str(e), token)
            last_err = f"ENTSO-E parse error: {type(e).__name__}: {safe_msg[:220]}"
            continue

    return empty, last_err or f"ENTSO-E returned no price data for {target_date} (tried A01 + A16)"


def _fetch_henex_dam_uncached(target_date: date, timeout: int = 20) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Raw Hellenic Energy Exchange (HEnEx/EnEx Group) Day-Ahead Market price
    fetch — a genuinely independent live source from ENTSO-E: separate
    organization, separate infrastructure (enexgroup.gr's own public
    document library), no authentication required. The URL pattern was
    obtained directly from trading@enexgroup.gr (their documented
    "automated downloading" endpoint, not reverse-engineered), and the
    result was verified against a live ENTSO-E pull for the same day:
    R²=1.0, zero difference across all 95 matched 15-min prices
    (2026-08-27), after correcting the MTU-numbering quirk documented below.

    File layout (per HEnEx's own PDF documentation): sheet
    "SPOT_Summary (SELL)", row 3 holds the MTU (Market Time Unit) column
    index (1..92/96/100 depending on DST), and the row labelled
    "Greece Mainland (15min MCP)" holds the 15-min clearing price per MTU.

    MTU numbering quirk (found empirically — NOT stated in HEnEx's
    documentation): despite row 3's date cell displaying midnight, MTU=1
    actually represents 01:00 Athens local time, not 00:00 (a legacy 1-24
    hour-labelling convention). The anchor below is deliberately local
    01:00 — confirmed to reproduce ENTSO-E's values exactly; anchoring at
    00:00 is off by exactly one hour.

    No Streamlit calls inside — safe to run from a worker thread.
    Returns (DataFrame[dt, price, _src], error_message_or_None).
    """
    empty = pd.DataFrame(columns=["dt", "price", "_src"])
    yyyymmdd = target_date.strftime("%Y%m%d")
    last_err = None

    for version in (1, 2):
        url = (f"https://www.enexgroup.gr/documents/20126/366820/"
              f"{yyyymmdd}_EL-DAM_ResultsSummary_EN_v{version:02d}.xlsx")
        try:
            r = requests.get(url, timeout=timeout, proxies=_get_proxies(), verify=True)
            if r.status_code == 404:
                last_err = f"HEnEx: no file published yet for {target_date} (v{version:02d})"
                continue
            if r.status_code != 200:
                last_err = f"HEnEx HTTP {r.status_code}"
                continue

            wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
            if len(rows) < 8:
                last_err = "HEnEx: workbook shorter than expected"
                continue

            row3 = rows[2]
            mtu_idx = [i for i, v in enumerate(row3) if isinstance(v, (int, float))]
            if not mtu_idx:
                last_err = "HEnEx: could not locate MTU index row"
                continue

            price_row = None
            for row in rows:
                label = row[0]
                if label and "mainland" in str(label).lower() and "mcp" in str(label).lower():
                    price_row = row
                    break
            if price_row is None:
                last_err = "HEnEx: could not locate 'Greece Mainland (...MCP)' row"
                continue

            prices = [price_row[i] for i in mtu_idx]
            if not all(isinstance(p, (int, float)) for p in prices):
                last_err = "HEnEx: non-numeric value in price row"
                continue

            n = len(prices)
            resolution_min = 15 if n > 30 else 60
            local_anchor = datetime(target_date.year, target_date.month, target_date.day,
                                    1, 0, tzinfo=ZoneInfo(PLANT_TZ))
            utc_anchor = local_anchor.astimezone(timezone.utc)
            timestamps = [utc_anchor + timedelta(minutes=resolution_min * i) for i in range(n)]

            df = pd.DataFrame({"dt": timestamps, "price": [float(p) for p in prices]})
            df["dt"] = pd.to_datetime(df["dt"], utc=True).dt.tz_convert(PLANT_TZ)
            df["_src"] = "HEnEx:DAM_ResultsSummary"
            _persist_dam_prices(df[["dt", "price"]], resolution_min=resolution_min, source="HENEX")
            return df[["dt", "price", "_src"]], None

        except Exception as e:
            last_err = f"HEnEx parse error: {type(e).__name__}: {str(e)[:220]}"
            continue

    return empty, last_err or f"HEnEx returned no data for {target_date}"


def _incomplete_day(day_df: pd.DataFrame) -> bool:
    """
    True if day_df has markedly fewer points than a full day at its own
    apparent resolution. Confirmed live (2026-03-29) that ENTSO-E can return
    HTTP 200 / no error while silently missing a multi-hour chunk of a day —
    "no error" alone isn't proof of a complete day.
    """
    if len(day_df) < 2:
        return True
    median_gap_min = day_df["dt"].diff().dt.total_seconds().median() / 60
    if median_gap_min <= 20:
        return len(day_df) < 90    # normal 15-min day; DST-short day has 92
    return len(day_df) < 22        # normal hourly day; DST-short day has 23


def _fetch_dam_daily_with_fallback(target_date: date, token: str) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Tries, in order: (1) ENTSO-E live, (2) HEnEx live — the Hellenic Energy
    Exchange, a genuinely independent organization/infrastructure, verified
    to match ENTSO-E exactly (see _fetch_henex_dam_uncached) — (3) the
    durable local dam_prices journal (whatever was last successfully
    fetched for this exact day, from either source).

    A successful ENTSO-E response is also checked for completeness, not
    just absence of an error — if the day is markedly short (see
    _incomplete_day), HEnEx is fetched once to fill in just the missing
    timestamps. ENTSO-E's own values are never overwritten, only genuine
    gaps are filled — and HEnEx is only fetched at all when a gap is
    actually suspected, so a normal complete day costs nothing extra.

    Still returns an error string even on a successful HEnEx/cache
    fallback (or a gap-fill), so callers can tell the difference from a
    fully live ENTSO-E value (dam_daily surfaces this via _entsoe_last_error).
    """
    df, err = _fetch_dam_daily_uncached(target_date, token)
    if not err:
        day_start = pd.Timestamp(target_date, tz=PLANT_TZ)
        day_end = day_start + timedelta(days=1)
        day_df = df[(df["dt"] >= day_start) & (df["dt"] < day_end)].sort_values("dt").reset_index(drop=True)

        if not _incomplete_day(day_df):
            return day_df, None

        henex_df, henex_err = _fetch_henex_dam_uncached(target_date)
        if henex_err or henex_df.empty:
            return day_df, (f"ENTSO-E response for {target_date} looks incomplete "
                            f"({len(day_df)} periods) and the HEnEx fill attempt failed: {henex_err}")

        have = set(day_df["dt"])
        fill = henex_df[~henex_df["dt"].isin(have)]
        if fill.empty:
            return day_df, None   # HEnEx had nothing extra — treat the ENTSO-E day as-is
        filled = pd.concat([day_df, fill], ignore_index=True).sort_values("dt").reset_index(drop=True)
        return filled, (f"ENTSO-E response for {target_date} was missing "
                        f"{len(fill)} period(s) — filled from HEnEx.")

    henex_df, henex_err = _fetch_henex_dam_uncached(target_date)
    if not henex_err:
        return henex_df, f"{err} — served from HEnEx (Hellenic Energy Exchange) instead."

    cached = _read_cached_dam_prices(target_date)
    if not cached.empty:
        return cached, f"{err}; {henex_err} — served {len(cached)} row(s) from local cache instead."
    return df, f"{err}; {henex_err}"


@st.cache_data(ttl=3600, show_spinner=False)
def dam_daily(target_date: date) -> pd.DataFrame:
    """
    Fetch 15-min Greek SMP from ENTSO-E Transparency Platform (A44, Day-Ahead).
    Returns DataFrame[dt, price, _src] or empty frame on failure.

    Requires ENTSO-E API token in secrets.toml:
        [entsoe]
        api_key = "your-token"

    Register free at: https://transparency.entsoe.eu

    For fetching several days at once, prefer dam_daily_batch() which runs the
    requests concurrently instead of one-by-one.
    """
    empty = pd.DataFrame(columns=["dt", "price", "_src"])
    token = _get_entsoe_token()

    if not token:
        st.session_state["_entsoe_last_error"] = (
            "No ENTSO-E API token found. Add [entsoe] api_key to secrets.toml. "
            "Register free at https://transparency.entsoe.eu"
        )
        return empty

    df, err = _fetch_dam_daily_with_fallback(target_date, token)
    if err:
        st.session_state["_entsoe_last_error"] = err
    else:
        st.session_state.pop("_entsoe_last_error", None)
    return df


def dam_daily_batch(dates: List[date]) -> Dict[date, pd.DataFrame]:
    """
    Fetch several days of 15-min DAM prices concurrently (ThreadPoolExecutor)
    instead of sequentially — cuts wall-clock time roughly by the worker count
    for multi-day pulls (e.g. the 7-day forecast tab).

    Results are memoised in session_state for the life of the session so
    repeated reruns of the same date range don't re-hit the network.
    """
    empty = pd.DataFrame(columns=["dt", "price", "_src"])
    cache: Dict[date, pd.DataFrame] = st.session_state.setdefault("_dam_daily_cache", {})
    token = _get_entsoe_token()

    result: Dict[date, pd.DataFrame] = {}
    missing = []
    for d in dict.fromkeys(dates):   # de-dupe, keep order
        if d in cache:
            result[d] = cache[d]
        else:
            missing.append(d)

    if not missing:
        return result

    if not token:
        st.session_state["_entsoe_last_error"] = (
            "No ENTSO-E API token found. Add [entsoe] api_key to secrets.toml. "
            "Register free at https://transparency.entsoe.eu"
        )
        for d in missing:
            cache[d] = empty
            result[d] = empty
        return result

    last_err = None
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(6, len(missing))) as ex:
        fut_map = {ex.submit(_fetch_dam_daily_with_fallback, d, token): d for d in missing}
        for fut in concurrent.futures.as_completed(fut_map):
            d = fut_map[fut]
            try:
                df, err = fut.result()
            except Exception as e:
                df, err = empty, _redact_token(str(e), token)
            if err:
                last_err = err
            cache[d] = df
            result[d] = df

    if last_err:
        st.session_state["_entsoe_last_error"] = last_err
    else:
        st.session_state.pop("_entsoe_last_error", None)
    return result


def _fetch_dam_monthly_avg_uncached(year: int, month: int, token: str,
                                    timeout: int = 30) -> Tuple[Optional[float], Optional[str]]:
    """
    Raw ENTSO-E monthly-average SMP fetch — no Streamlit calls inside, so it's
    safe to run from a worker thread (used by dam_monthly_avg / _batch).
    Returns (avg_price_or_None, error_message_or_None).
    """
    last_day     = calendar.monthrange(year, month)[1]
    period_start = datetime(year, month, 1, 0, 0, tzinfo=timezone.utc)
    period_end   = datetime(year, month, last_day, 23, 0, tzinfo=timezone.utc)

    import xml.etree.ElementTree as ET

    last_err = None
    for process_type in ["A01", "A16"]:
        params = {
            "securityToken": token,
            "documentType":  "A44",
            "processType":   process_type,
            "in_Domain":     ENTSOE_ZONE,
            "out_Domain":    ENTSOE_ZONE,
            "periodStart":   period_start.strftime("%Y%m%d%H%M"),
            "periodEnd":     period_end.strftime("%Y%m%d%H%M"),
        }
        try:
            r = _ENTSOE_SESSION.get(ENTSOE_API, params=params, timeout=timeout,
                                    proxies=_get_proxies(), verify=False)
            if r.status_code != 200:
                last_err = f"ENTSO-E HTTP {r.status_code}"
                continue

            root      = ET.fromstring(r.content)
            root_tag  = root.tag
            actual_ns = root_tag.split("}")[0].lstrip("{") if "}" in root_tag else ""

            if "acknowledgement" in actual_ns.lower():
                continue

            ns     = {"ns": actual_ns} if actual_ns else {}
            prefix = "ns:" if actual_ns else ""
            prices = []
            for pt in root.findall(f".//{prefix}Point", ns):
                v = pt.findtext(f"{prefix}price.amount", namespaces=ns)
                if v:
                    fv = float(v)
                    if fv > 0:
                        prices.append(fv)

            if prices:
                return float(np.mean(prices)), None

        except Exception as e:
            safe_msg = _redact_token(str(e), token)
            last_err = f"ENTSO-E monthly error: {type(e).__name__}: {safe_msg[:220]}"
            continue

    return None, last_err


@st.cache_data(ttl=7200, show_spinner=False)
def dam_monthly_avg(year: int, month: int) -> Optional[float]:
    """
    Monthly average Greek SMP from ENTSO-E (A44, Day-Ahead).
    Falls back to SQLite price cache if API token missing or unreachable.

    For fetching several months at once, prefer dam_monthly_avg_batch() which
    runs the requests concurrently instead of one-by-one.
    """
    ym    = f"{year}-{month:02d}"
    token = _get_entsoe_token()

    if not token:
        return _get_cached_dam_price(ym)

    avg, err = _fetch_dam_monthly_avg_uncached(year, month, token)
    if avg is not None:
        _cache_dam_price(ym, avg)
        return avg
    if err:
        st.session_state["_entsoe_last_error"] = err
    return _get_cached_dam_price(ym)


def dam_monthly_avg_batch(pairs: List[Tuple[int, int]]) -> Dict[Tuple[int, int], Optional[float]]:
    """
    Fetch several (year, month) DAM averages concurrently (ThreadPoolExecutor)
    instead of sequentially — cuts wall-clock time roughly by the worker count
    for multi-month pulls (e.g. the Monthly/Financial tabs' YTD price lookups).

    Results are memoised in session_state for the life of the session so
    repeated reruns of the same period don't re-hit the network.
    """
    cache: Dict[Tuple[int, int], Optional[float]] = st.session_state.setdefault(
        "_dam_monthly_cache", {})
    token = _get_entsoe_token()

    result: Dict[Tuple[int, int], Optional[float]] = {}
    missing = []
    for p in dict.fromkeys(pairs):   # de-dupe, keep order
        if p in cache:
            result[p] = cache[p]
        else:
            missing.append(p)

    if not missing:
        return result

    if not token:
        for p in missing:
            v = _get_cached_dam_price(f"{p[0]}-{p[1]:02d}")
            cache[p] = v
            result[p] = v
        return result

    last_err = None
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(6, len(missing))) as ex:
        fut_map = {ex.submit(_fetch_dam_monthly_avg_uncached, y, m, token): (y, m)
                   for (y, m) in missing}
        for fut in concurrent.futures.as_completed(fut_map):
            p = fut_map[fut]
            try:
                avg, err = fut.result()
            except Exception as e:
                avg, err = None, _redact_token(str(e), token)
            ym = f"{p[0]}-{p[1]:02d}"
            if avg is not None:
                _cache_dam_price(ym, avg)
            else:
                avg = _get_cached_dam_price(ym)
                if err:
                    last_err = err
            cache[p] = avg
            result[p] = avg

    if last_err:
        st.session_state["_entsoe_last_error"] = last_err
    return result

# ── Daily production-weighted (capture-price) revenue ──────────────────────
# Same calculation the Intraday tab shows for one day — 5-min/hourly production
# matched against 15-min DAM prices — extracted so it can be computed in bulk
# for the Monthly tab's revenue column and persisted to daily_revenue so
# repeat visits (and other tabs) reuse it instead of re-fetching.
def _compute_day_revenue_from_frames(prod_json: dict, df_dam: pd.DataFrame) -> Optional[dict]:
    """Pure computation, no network/Streamlit calls — safe to call from anywhere."""
    raw = prod_json.get("data", []) if prod_json else []
    src = prod_json.get("_src", "hour") if prod_json else "hour"
    if raw and isinstance(raw, list) and "kpiList" in (raw[0] if isinstance(raw[0], dict) else {}):
        raw = raw[0]["kpiList"]
    df_p = _norm(raw)
    if df_p.empty or df_dam.empty or "price" not in df_dam.columns:
        return None

    tc = next((c for c in df_p.columns if "time" in c.lower() or "collect" in c.lower()), None)
    yc = next((c for c in df_p.columns if "inverterYield" in c or "activePower" in c
              or "day_cap" in c or "power" in c.lower()), None)
    if not tc or not yc:
        return None

    df_p["dt"] = pd.to_datetime(df_p[tc], unit="ms", utc=True, errors="coerce").dt.tz_convert(PLANT_TZ)
    df_p[yc]   = pd.to_numeric(df_p[yc], errors="coerce")
    df_p = df_p.dropna(subset=["dt", yc]).sort_values("dt").reset_index(drop=True)
    if len(df_p) < 2:
        return None

    median_gap   = df_p["dt"].diff().dropna().dt.total_seconds().median()
    interval_min = int(round(median_gap / 60)) if pd.notna(median_gap) and median_gap > 0 else 60
    if src == "hour" and len(df_p) <= 25:
        # Mirrors the Intraday tab: hourly-fallback data can have irregular
        # gaps (e.g. missing night-time entries) that skew the median-gap
        # estimate upward, which directly inflates kWh — and revenue with it.
        interval_min = 60

    if interval_min > 15:
        # Hourly-fallback data is coarser than the 15-min buckets used
        # everywhere else (DAM prices, revenue_15min). Rather than dumping a
        # whole hour's energy into a single 15-min bucket (spiking it to
        # interval_min/15x the true power and leaving the other three
        # empty), expand each hourly reading into its four constituent
        # 15-min slots at the same kW value — the best available estimate
        # for each quarter given only an hourly average. Total daily kWh is
        # unaffected: 4 quarters x (V kW * 0.25h) == 1 hour x (V kW * 1h).
        df_p = pd.concat([
            df_p.assign(dt=df_p["dt"] + pd.Timedelta(minutes=m))
            for m in range(0, interval_min, 15)
        ], ignore_index=True)
        interval_min = 15

    df_p["bucket"] = df_p["dt"].dt.floor("15min")
    bucket_prod = (df_p.groupby("bucket")[yc].sum().reset_index()
                  .rename(columns={"bucket": "dt", yc: "kw_sum"}))
    # Both sides are already tz-aware in PLANT_TZ, but merge_asof requires an
    # exact dtype match, not just equivalent UTC instants — and depending on
    # how each side's timestamps were built upstream (fixed-offset arithmetic
    # vs a named zone), pandas can hand back "Europe/Athens" on one side and
    # a fixed "UTC+03:00"/"UTC+02:00" offset on the other (seen live on
    # Streamlit Cloud's pandas/Python build), which raises MergeError even
    # though the values themselves agree. Normalizing both to UTC — the one
    # representation that's always canonical regardless of construction path
    # — sidesteps the whole class of mismatch; convert back to PLANT_TZ after.
    left  = bucket_prod.assign(dt=bucket_prod["dt"].dt.tz_convert("UTC")).sort_values("dt")
    right = (df_dam[["dt", "price"]].assign(dt=df_dam["dt"].dt.tz_convert("UTC"))
            .sort_values("dt"))
    dr = pd.merge_asof(left, right, on="dt", direction="backward",
                       tolerance=pd.Timedelta("16min"))
    dr["dt"] = dr["dt"].dt.tz_convert(PLANT_TZ)
    dr["kwh"] = dr["kw_sum"] * (interval_min / 60)
    dr["rev"] = dr["kwh"] / 1000 * dr["price"]

    kwh = float(df_p[yc].sum() * interval_min / 60)
    if kwh <= 0:
        return None
    revenue = float(dr["rev"].sum(skipna=True))
    return {"kwh": kwh, "revenue_eur": revenue, "avg_price": float(df_dam["price"].mean()),
           "buckets": dr[["dt", "kwh", "price", "rev"]].dropna(subset=["price"])}


def _cached_revenue_days(dates: List[date]) -> Dict[str, dict]:
    """Read whichever of `dates` already have a daily_revenue row."""
    if not dates:
        return {}
    conn = _get_db()
    placeholders = ",".join("?" * len(dates))
    rows = conn.execute(
        f"SELECT day,kwh,revenue_eur,avg_price FROM daily_revenue WHERE day IN ({placeholders})",
        [str(d) for d in dates]).fetchall()
    conn.close()
    return {row[0]: {"kwh": row[1], "revenue_eur": row[2], "avg_price": row[3]} for row in rows}


def _persist_day_revenue(d: date, r: dict) -> None:
    """
    Write one day's revenue result — as produced by
    _compute_day_revenue_from_frames — to daily_revenue + revenue_15min.
    This is the ONLY place either table gets written to: both
    compute_daily_revenue_batch (Monthly tab) and the Intraday tab's save
    action call this, so a day saved from either tab is calculated the same
    way and lands in the same storage, instead of two independently
    maintained copies that can silently drift apart.
    """
    fetched_ts = datetime.now(timezone.utc).isoformat()
    conn = _get_db()
    conn.execute("""INSERT OR REPLACE INTO daily_revenue
        (day,kwh,revenue_eur,avg_price,fetched_ts) VALUES (?,?,?,?,?)""",
        (str(d), r["kwh"], r["revenue_eur"], r["avg_price"], fetched_ts))
    buckets = r.get("buckets")
    if buckets is not None and not buckets.empty:
        conn.executemany("""INSERT OR REPLACE INTO revenue_15min
            (dt,day,kwh,price_eur_mwh,revenue_eur,fetched_ts) VALUES (?,?,?,?,?,?)""",
            [(b["dt"].isoformat(), str(d), float(b["kwh"]), float(b["price"]),
             float(b["rev"]), fetched_ts) for _, b in buckets.iterrows()])
    conn.commit(); conn.close()


def compute_daily_revenue_batch(base_url, sid, dates: List[date], xsrf, verify,
                                request_gap_seconds: float = 30.0,
                                on_progress=None) -> Tuple[Dict[date, dict], Dict[date, str]]:
    """
    Compute (or reuse cached) 15-min production-weighted revenue for several
    days. Days already in the daily_revenue SQLite table are reused for free.

    Missing days are fetched from Huawei strictly ONE AT A TIME with a
    mandatory pause between requests — confirmed by direct reproduction that
    Huawei's KPI endpoints (getKpiStation5min/Hour) enforce an account-wide
    rate limit stricter than a short burst window: a single call succeeds,
    then every subsequent call returns failCode=407 for 2+ minutes straight
    regardless of concurrency or retries. Concurrency here would only get the
    account rate-limited faster, not finish faster.

    ENTSO-E price lookups are NOT subject to this and are still fetched via
    the existing parallel dam_daily_batch — this constraint is specific to
    Huawei's account.

    Each day is persisted to daily_revenue/revenue_15min as soon as it's
    computed (not batched to the end), so a run that's interrupted partway
    through — very possible given how long this can take at a safe pace —
    doesn't lose the days it already finished.

    on_progress(day, ok, reason), if given, is called after every day
    (cached or freshly fetched) so a caller can render live progress.

    Returns (result, fail_reasons) — fail_reasons maps any day that couldn't
    be computed to the raw failCode/message Huawei/ENTSO-E returned, so the
    caller can tell a real rate-limit/auth failure apart from "no data".
    """
    if not dates:
        return {}, {}
    cached = _cached_revenue_days(dates)
    result  = {d: cached[str(d)] for d in dates if str(d) in cached}
    missing = [d for d in dates if str(d) not in cached]
    for d in result:
        if on_progress: on_progress(d, True, None)
    if not missing:
        return result, {}

    dam_raw = dam_daily_batch(missing)
    fail_reasons = {}

    for idx, d in enumerate(missing):
        pj = api_15min(base_url, sid, d, xsrf, verify) or {}
        r = _compute_day_revenue_from_frames(pj, dam_raw.get(d, pd.DataFrame()))

        if r:
            result[d] = {"kwh": r["kwh"], "revenue_eur": r["revenue_eur"],
                        "avg_price": r["avg_price"]}
            _persist_day_revenue(d, r)
            reason = None
        else:
            fc = pj.get("failCode")
            msg = pj.get("message") or pj.get("msg")
            if fc is not None or msg:
                reason = f"failCode={fc}" + (f" — {msg}" if msg else "")
            elif dam_raw.get(d, pd.DataFrame()).empty:
                reason = "no ENTSO-E price data"
            else:
                reason = "no production data"
            fail_reasons[d] = reason

        if on_progress: on_progress(d, r is not None, reason)
        if idx < len(missing) - 1:
            time.sleep(request_gap_seconds)

    return result, fail_reasons

# ─────────────────────────────────────────────────────────────────────────────
# ALERT ENGINE — run after every data load
# ─────────────────────────────────────────────────────────────────────────────
def run_alert_checks(df_bench: pd.DataFrame):
    """Evaluate rule-based alerts on the monthly benchmark dataframe and persist."""
    if df_bench.empty: return
    recent = df_bench.dropna(subset=["PR"]).tail(3)
    if len(recent)==3 and (recent["PR"] < PR_ALERT_RED).all():
        add_alert("Critical","PR",
            f"PR below {PR_ALERT_RED:.0%} for 3 consecutive months "
            f"(last: {recent['PR'].iloc[-1]:.1%})")
    last = df_bench.dropna(subset=["Energy_kWh","Expected_kWh"]).iloc[-1] \
           if not df_bench.empty else None
    if last is not None and last["Expected_kWh"]>0:
        ratio = last["Energy_kWh"] / last["Expected_kWh"]
        if ratio < PROD_VS_EXP_WARN:
            add_alert("Warning","Production",
                f"Last month actual production {ratio:.0%} of PVGIS expected "
                f"({last['Energy_kWh']:,.0f} vs {last['Expected_kWh']:,.0f} kWh)")

# ─────────────────────────────────────────────────────────────────────────────
# WEATHER FORECAST (Open-Meteo — free, no key)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=3600, show_spinner=False)
def fetch_forecast() -> pd.DataFrame:
    """7-day hourly GHI + T_amb from Open-Meteo for Thessaloniki."""
    try:
        r = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params={"latitude":40.694,"longitude":22.990,
                    "hourly":"shortwave_radiation,temperature_2m",
                    "timezone":"Europe/Athens","forecast_days":7},
            timeout=15,
            proxies=_get_proxies(),
            verify=False)
        if r.status_code!=200:
            return pd.DataFrame()
        j = r.json()
        df = pd.DataFrame({"dt": pd.to_datetime(j["hourly"]["time"]),
                           "GHI_Wm2": j["hourly"]["shortwave_radiation"],
                           "T_amb":   j["hourly"]["temperature_2m"]})
        df["dt"] = df["dt"].dt.tz_localize("Europe/Athens", ambiguous="NaT",
                                           nonexistent="NaT")
        df["GTI_Wm2"] = df["GHI_Wm2"] * 1.15
        t_cell = df["T_amb"] + (NOCT-20)/800 * df["GTI_Wm2"]
        df["Yield_kWh"] = (df["GTI_Wm2"]/1000 * PLANT_PEAK_KW
                           * (1 + GAMMA*(t_cell-25)) * 0.78).clip(lower=0)
        return df.dropna()
    except Exception as e:
        st.session_state["_forecast_error"] = str(e)
        return pd.DataFrame()

# ─────────────────────────────────────────────────────────────────────────────
# DEGRADATION FIT
# ─────────────────────────────────────────────────────────────────────────────
def fit_degradation(df_monthly: pd.DataFrame) -> Optional[dict]:
    """
    Linear regression on annual specific yield (kWh/kWp) vs year.
    Returns slope (%/year), intercept, r², projection dict.
    """
    try:
        from scipy.stats import linregress
        df_a = (df_monthly.groupby("Year")["Energy_kWh"].sum()
                .reset_index())
        df_a["Specific_Yield"] = df_a["Energy_kWh"] / PLANT_PEAK_KW
        if len(df_a) < 2: return None
        sl,ic,r,p,se = linregress(df_a["Year"], df_a["Specific_Yield"])
        pct_yr = sl / ic * 100
        return {"slope":sl,"intercept":ic,"r2":r**2,"pct_per_year":pct_yr,
                "data":df_a}
    except: return None

# ─────────────────────────────────────────────────────────────────────────────
# ML ANOMALY DETECTION (Isolation Forest)
# ─────────────────────────────────────────────────────────────────────────────
def run_anomaly_detection(df_all: pd.DataFrame) -> pd.DataFrame:
    """
    Isolation Forest on monthly aggregated telemetry features.
    Returns df with anomaly_score column (-1 = anomaly).
    """
    try:
        from sklearn.ensemble import IsolationForest
        feat_cols = []
        for c in df_all.columns:
            cl = c.lower()
            if any(k in cl for k in ["efficiency","power_factor","temperature",
                                     "active_power","elec_freq"]):
                feat_cols.append(c)
        if not feat_cols: return pd.DataFrame()
        df_f = df_all[feat_cols].apply(pd.to_numeric, errors="coerce").dropna()
        if len(df_f) < 10: return pd.DataFrame()
        clf = IsolationForest(contamination=0.1, random_state=42)
        df_f["anomaly"] = clf.fit_predict(df_f)
        df_f["score"]   = clf.score_samples(df_f[feat_cols])
        return df_f
    except: return pd.DataFrame()

# ─────────────────────────────────────────────────────────────────────────────
# PDF REPORT
# ─────────────────────────────────────────────────────────────────────────────
def build_pdf_report(merged, df_bench, alert_rows, year) -> Optional[bytes]:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        H1 = ParagraphStyle("h1",parent=styles["Heading1"],
                            fontSize=18,textColor=colors.HexColor("#f0b429"))
        H2 = ParagraphStyle("h2",parent=styles["Heading2"],
                            fontSize=12,textColor=colors.HexColor("#3ecfcf"))
        BODY = ParagraphStyle("body",parent=styles["Normal"],fontSize=9)
        story=[]
        story.append(Paragraph(f"☀ FusionSolar APM — Monthly Report {year}", H1))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", BODY))
        story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor("#334155")))
        story.append(Spacer(1,0.4*cm))

        # KPI table
        story.append(Paragraph("Performance Summary", H2))
        ytd_kwh  = merged["Energy_kWh"].sum(skipna=True)
        ytd_budg = merged["Budget_kWh"].sum()
        ach = ytd_kwh/ytd_budg*100 if ytd_budg else 0
        avg_pr = df_bench["PR"].mean() if "PR" in df_bench.columns else np.nan
        data=[["KPI","Value"],
              ["YTD Production (kWh)", f"{ytd_kwh:,.0f}"],
              ["YTD Budget (kWh)", f"{ytd_budg:,.0f}"],
              ["Achievement", f"{ach:.1f}%"],
              ["Avg PR", f"{avg_pr:.2f}" if not np.isnan(avg_pr) else "—"]]
        t=Table(data,colWidths=[9*cm,7*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1f2333")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#f0b429")),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#334155")),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),
             [colors.HexColor("#161b22"),colors.HexColor("#0e1117")]),
            ("TEXTCOLOR",(0,1),(-1,-1),colors.HexColor("#e2e8f0"))]))
        story.append(t)
        story.append(Spacer(1,0.5*cm))

        # Monthly detail
        story.append(Paragraph("Monthly Energy vs Budget", H2))
        hdr=["Month","Budget kWh","Actual kWh","Delta kWh","Achievement"]
        rows_=[hdr]
        for _,row in merged.iterrows():
            rows_.append([row["Month"],
                f"{row['Budget_kWh']:,.0f}",
                f"{row['Energy_kWh']:,.0f}" if pd.notna(row.get('Energy_kWh')) else "—",
                f"{row.get('Delta_kWh',np.nan):+,.0f}" if pd.notna(row.get('Delta_kWh')) else "—",
                f"{row.get('Achievement_%',np.nan):.1f}%" if pd.notna(row.get('Achievement_%')) else "—"])
        t2=Table(rows_,colWidths=[3*cm,3.5*cm,3.5*cm,3.5*cm,3*cm])
        t2.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1f2333")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#f0b429")),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#334155")),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),
             [colors.HexColor("#161b22"),colors.HexColor("#0e1117")]),
            ("TEXTCOLOR",(0,1),(-1,-1),colors.HexColor("#e2e8f0"))]))
        story.append(t2)
        story.append(Spacer(1,0.5*cm))

        # Alerts
        if alert_rows:
            story.append(Paragraph("Open Alerts", H2))
            ah=[["Timestamp","Severity","Category","Message"]]
            for a in alert_rows[:10]:
                ah.append([a[1],a[2],a[3],a[4][:80]])
            ta=Table(ah,colWidths=[3.5*cm,2.5*cm,3*cm,8*cm])
            ta.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1f2333")),
                ("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#fb7185")),
                ("FONTSIZE",(0,0),(-1,-1),7.5),
                ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#334155")),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),
                 [colors.HexColor("#161b22"),colors.HexColor("#0e1117")]),
                ("TEXTCOLOR",(0,1),(-1,-1),colors.HexColor("#e2e8f0"))]))
            story.append(ta)

        doc.build(story)
        return buf.getvalue()
    except Exception as e:
        st.warning(f"PDF generation failed: {e}. Install reportlab: pip install reportlab")
        return None

# ─────────────────────────────────────────────────────────────────────────────
# ██████████████████  MAIN UI  ████████████████████████████████████████████████
# ─────────────────────────────────────────────────────────────────────────────

# ── Sidebar ──
with st.sidebar:
    st.title("☀️ FusionSolar APM")
    year_input = int(st.number_input("Analysis Year",
                    min_value=PLANT_START_YEAR, max_value=2030, value=date.today().year))
    st.caption(f"Capacity: {PLANT_PEAK_KW:.0f} kWp | COD: {PLANT_START_YEAR}")
    st.caption("📍 Thessaloniki (PVGIS-SARAH3)")
    st.divider()

    st.subheader("Finance Inputs")
    annual_debt  = st.number_input("Annual Debt Service (€)", value=ANNUAL_DEBT_SVC,
                                   step=5000.0, format="%.0f")
    fixed_opex   = st.number_input("Fixed OPEX/year (€)", value=FIXED_OPEX,
                                   step=1000.0, format="%.0f")
    var_opex     = st.number_input("Variable OPEX (€/MWh)", value=VAR_OPEX_PER_MWH,
                                   step=0.1, format="%.2f")
    ref_pr       = st.slider("Reference PR (irradiance model)", 0.60, 0.95, 0.78, 0.01)
    st.divider()

    if st.button("🔄 Reset Login"):
        st.session_state.pop("hw_client",None)
        st.session_state.pop("hw_stations",None)
        st.rerun()

# ── Tabs ──
(t_score, t_monthly, t_intraday, t_hist,
 t_loss, t_financial, t_forecast,
 t_health, t_failure, t_opex, t_alerts, t_events, t_ipto) = st.tabs([
    "🎯 Scorecard",
    "📊 Monthly",
    "📈 Intraday",
    "📉 Historical",
    "🔻 Loss Cascade",
    "💰 Financial",
    "🔮 Forecast",
    "🛠️ Health",
    "🩺 Failure Analytics",
    "🧾 OPEX",
    "🔔 Alerts",
    "🚨 Events",
    "📡 ENTSO-E Diagnostics",
])

# ─────────────────────────────────────────────────────────────────────────────
# TAB: SCORECARD  (Tier 1 — Executive KPI Strip)
# ─────────────────────────────────────────────────────────────────────────────
with t_score:
    st.header("Executive Performance Scorecard")
    st.caption("Live plant status — reloads automatically on page refresh.")

    def _load_scorecard(year_input):
        client, stations = ensure_client()
        if not client or not stations:
            return
        sid = stations[0].get("stationCode") or stations[0].get("plantCode")
        with st.spinner("Loading scorecard…"):
            yr_data = api_monthly_years(client.base_url, sid,
                                        [year_input, year_input-1],
                                        client.xsrf, client.verify_ssl)
            df_yr = yr_data[year_input]
            df_py = yr_data[year_input-1]

            if not df_yr.empty:
                tc, ec = _resolve(df_yr)
                if tc and ec:
                    df_yr["dt"]  = pd.to_datetime(df_yr[tc],unit="ms",utc=True).dt.tz_convert(PLANT_TZ)
                    df_yr["kWh"] = pd.to_numeric(df_yr[ec],errors="coerce")
                    df_yr["m"]   = df_yr["dt"].dt.month

                    ytd_kwh  = df_yr["kWh"].sum(skipna=True)
                    bdf      = get_budget(year_input)
                    ytd_budg = bdf["Budget_kWh"].sum()
                    ach_pct  = ytd_kwh/ytd_budg*100 if ytd_budg else 0

                    # PR from PVGIS
                    gti_ytd  = sum(PVGIS_GTI[m] for m in df_yr["m"])
                    pr_ytd   = ytd_kwh/(gti_ytd*PLANT_PEAK_KW) if gti_ytd else np.nan

                    # WCPR
                    wcpr_vals=[]
                    for _,row in df_yr.iterrows():
                        g = PVGIS_GTI.get(int(row["m"]),0)
                        ta= T_AMB.get(int(row["m"]),20)
                        if g>0: wcpr_vals.append(wcpr(row["kWh"],g,ta))
                    wcpr_avg = float(np.nanmean(wcpr_vals)) if wcpr_vals else np.nan

                    # YoY delta
                    py_kwh = np.nan
                    if not df_py.empty:
                        tc2,ec2 = _resolve(df_py)
                        if tc2 and ec2:
                            py_kwh = pd.to_numeric(df_py[ec2],errors="coerce").sum()

                    # Alerts
                    conn = _get_db()
                    n_alerts = conn.execute(
                        "SELECT COUNT(*) FROM alerts WHERE status='Open'").fetchone()[0]
                    conn.close()

                    # RAG colour for PR
                    def _rag(v):
                        if np.isnan(v): return "⚫"
                        if v>=PR_ALERT_AMBER: return "🟢"
                        if v>=PR_ALERT_RED:   return "🟡"
                        return "🔴"

                    # Scorecard row
                    c1,c2,c3,c4,c5,c6 = st.columns(6)
                    c1.metric("YTD Production",f"{ytd_kwh/1e6:.3f} GWh",
                              delta=f"{(ytd_kwh-py_kwh)/1e3:.0f} MWh vs LY" if not np.isnan(py_kwh) else None)
                    c2.metric("vs Budget",f"{ach_pct:.1f}%",
                              delta=f"{ytd_kwh-ytd_budg:+,.0f} kWh")
                    c3.metric(f"PR {_rag(pr_ytd)}",f"{pr_ytd:.3f}" if not np.isnan(pr_ytd) else "—")
                    c4.metric("WCPR ⭐",f"{wcpr_avg:.3f}" if not np.isnan(wcpr_avg) else "—",
                              help="Weather-corrected PR — IEC 61724")
                    c5.metric("Specific Yield",
                              f"{ytd_kwh/PLANT_PEAK_KW:.0f} kWh/kWp")
                    c6.metric("Open Alerts",str(n_alerts),
                              delta="⚠️" if n_alerts>0 else None,
                              delta_color="inverse" if n_alerts>0 else "normal")

                    st.divider()

                    # Monthly bar chart with PR overlay
                    fig = make_subplots(specs=[[{"secondary_y":True}]])
                    bdf2 = get_budget(year_input)
                    fig.add_bar(x=MONTH_LABELS, y=bdf2["Budget_kWh"],
                                name="Budget", marker_color="#334155",
                                secondary_y=False)
                    fig.add_bar(x=[MONTH_LABELS[r["m"]-1] for _,r in df_yr.iterrows()],
                                y=df_yr["kWh"],
                                name="Actual", marker_color="#f0b429",
                                secondary_y=False)
                    # PR per month
                    pr_vals = [wcpr(row["kWh"],PVGIS_GTI.get(int(row["m"]),1),
                                    T_AMB.get(int(row["m"]),20))
                               for _,row in df_yr.iterrows()]
                    fig.add_scatter(
                        x=[MONTH_LABELS[r["m"]-1] for _,r in df_yr.iterrows()],
                        y=pr_vals, mode="lines+markers", name="WCPR",
                        line=dict(color="#3ecfcf",width=2),
                        marker=dict(size=7), secondary_y=True)
                    fig.add_hline(y=PR_ALERT_AMBER, line_dash="dash",
                                  line_color="#4ade80",
                                  annotation_text=f"Target {PR_ALERT_AMBER}",
                                  secondary_y=True)
                    fig.add_hline(y=PR_ALERT_RED, line_dash="dot",
                                  line_color="#ff5f5f",
                                  annotation_text=f"Alert {PR_ALERT_RED}",
                                  secondary_y=True)
                    _dual_layout(fig,
                        title=f"{year_input} Monthly Production & WCPR",
                        left_title="Energy (kWh)", right_title="WCPR",
                        left_color="#f0b429", right_color="#3ecfcf")
                    fig.update_layout(barmode="group")
                    st.plotly_chart(fig, use_container_width=True)

                    st.session_state["scorecard_merged"] = bdf2.merge(
                        df_yr[["m","kWh"]].rename(columns={"m":"MonthNum","kWh":"Energy_kWh"}),
                        on="MonthNum", how="left")
                    st.session_state["scorecard_year"] = year_input

    _load_scorecard(year_input)

# ─────────────────────────────────────────────────────────────────────────────
# TAB: MONTHLY  (production vs budget + revenue)
# ─────────────────────────────────────────────────────────────────────────────
with t_monthly:
    st.header("Monthly Energy vs Budget")
    today = date.today()

    # Only the Huawei monthly-KPI pull is gated behind this — everything
    # else below (revenue button, charts, table) must run on EVERY rerun,
    # otherwise clicking any button on this tab triggers a rerun where
    # _run_monthly is False and the whole section disappears.
    _run_monthly = (st.button("🔄 Refresh Monthly", key="btn_monthly")
                    or "monthly_raw" not in st.session_state
                    or st.session_state.get("monthly_year") != year_input)

    if _run_monthly:
        client, stations = ensure_client()
        if not client or not stations:
            st.error("❌ FusionSolar connection failed — check secrets.toml.")
            st.stop()
        sid = stations[0].get("stationCode") or stations[0].get("plantCode")

        with st.spinner("Fetching…"):
            yr_data = api_monthly_years(client.base_url, sid,
                                        [year_input, year_input-1],
                                        client.xsrf, client.verify_ssl)
            df_raw  = yr_data[year_input]
            df_prev = yr_data[year_input-1]

        if df_raw.empty:
            st.warning("No monthly data."); st.stop()
        tc,ec = _resolve(df_raw)
        if not tc or not ec:
            st.error(f"Columns: {list(df_raw.columns)}"); st.stop()

        df_raw["dt"]  = pd.to_datetime(df_raw[tc],unit="ms",utc=True).dt.tz_convert(PLANT_TZ)
        df_raw["m"]   = df_raw["dt"].dt.month
        df_raw["kWh"] = pd.to_numeric(df_raw[ec],errors="coerce")
        bdf = get_budget(year_input)

        # Prorate budget for the current in-progress month
        if year_input == today.year:
            cur_m      = today.month
            days_so_far = today.day
            days_total  = calendar.monthrange(today.year, cur_m)[1]
            mask = bdf["MonthNum"] == cur_m
            bdf.loc[mask, "Budget_kWh"] = (
                bdf.loc[mask, "Budget_kWh"] * days_so_far / days_total
            ).round(0)
        merged = bdf.merge(df_raw[["m","kWh"]].rename(
            columns={"m":"MonthNum","kWh":"Energy_kWh"}),
            on="MonthNum", how="left")
        merged["Delta_kWh"]     = merged["Energy_kWh"]-merged["Budget_kWh"]
        merged["Achievement_%"] = (
            merged["Energy_kWh"].astype(float)
            / merged["Budget_kWh"].astype(float).where(merged["Budget_kWh"] != 0)
            * 100
        ).round(1)

        # Rolling 12-mo source frames
        frames=[]
        for df_y,yr_l in [(df_prev,year_input-1),(df_raw,year_input)]:
            if df_y.empty: continue
            t2,e2=_resolve(df_y)
            if t2 and e2:
                tmp=df_y[[t2,e2]].copy(); tmp.columns=["ts","kWh"]
                tmp["dt"]=pd.to_datetime(tmp["ts"],unit="ms",utc=True).dt.tz_convert(PLANT_TZ)
                tmp["m"]=tmp["dt"].dt.month; tmp["yr"]=yr_l
                tmp["kWh"]=pd.to_numeric(tmp["kWh"],errors="coerce")
                frames.append(tmp)

        st.session_state["monthly_raw"]    = merged     # pre-revenue, cached
        st.session_state["monthly_frames"] = frames
        st.session_state["monthly_year"]   = year_input

    if "monthly_raw" not in st.session_state:
        st.info("Loading…"); st.stop()

    merged = st.session_state["monthly_raw"].copy()
    frames = st.session_state["monthly_frames"]

    # Revenue — single method: Σ (15-min ENTSO-E price × generation in that
    # 15-min period), summed per day then rolled up to months. No
    # average-price shortcut, no fallback — a month is either computed this
    # way or shows as not-yet-computed. Runs every rerun (not gated behind
    # _run_monthly) so the button below actually works.
    elapsed_days=[date(year_input,m,d)
                 for m in range(1,13)
                 for d in range(1,calendar.monthrange(year_input,m)[1]+1)
                 if date(year_input,m,d)<today]
    # SQLite (daily_revenue) is the single source of truth for what's been
    # computed — NOT session_state. A day written to the DB last week, last
    # session, or by any other process (a backfill script, another browser
    # tab) must show up here immediately. Previously this table's rollup was
    # built from a session_state dict that only accumulated whatever THIS
    # browser session had explicitly fetched, so a fresh session showed a
    # fully-cached month (e.g. January, 31/31 in the DB) as mostly missing —
    # that's the "8/31" bug.
    cached_days=_cached_revenue_days(elapsed_days)
    new_days=[d for d in elapsed_days if str(d) not in cached_days]

    st.caption(
        "Revenue = Σ (15-min ENTSO-E price × generation in that period), summed per day. "
        f"{len(cached_days)}/{len(elapsed_days)} elapsed day(s) already computed & cached."
    )
    if new_days:
        pc1,pc2,pc3=st.columns([1,1,1.4])
        with pc1:
            pace=st.number_input("Seconds between requests",min_value=5,max_value=300,
                                 value=30,step=5,key="rev_pace",
                                 help="Huawei's KPI API rate-limits hard (confirmed: even "
                                      "30s spacing still failed ~29/30 requests in testing). "
                                      "Lower this only if you've confirmed a shorter gap "
                                      "reliably works for your account.")
        with pc2:
            day_cap=st.number_input("Days this run",min_value=1,
                                    max_value=len(new_days),
                                    value=min(20,len(new_days)),step=5,key="rev_cap")
        with pc3:
            est_min=day_cap*pace/60
            st.caption(f"~{est_min:.0f} min for {day_cap} day(s) "
                      f"({len(new_days)} total not yet computed).")
        compute_15min=st.button("📊 Compute Revenue",key="btn_15min_rev")
    else:
        compute_15min=False
        st.caption("✅ All elapsed days already computed.")

    if compute_15min and new_days:
        client, stations = ensure_client()
        if not client or not stations:
            st.error("❌ FusionSolar connection failed — check secrets.toml.")
            st.stop()
        sid = stations[0].get("stationCode") or stations[0].get("plantCode")
        todo=new_days[:int(day_cap)]
        prog=st.progress(0.0,f"0/{len(todo)} days…")
        status=st.empty()
        counts={"done":0}
        def _on_progress(d,ok,reason):
            counts["done"]+=1
            prog.progress(counts["done"]/len(todo),f"{counts['done']}/{len(todo)} days")
            status.caption(("✅ " if ok else "❌ ") + str(d) + (f" — {reason}" if reason else ""))
        # Fetched strictly one day at a time inside compute_daily_revenue_batch
        # (see its docstring) and persisted to SQLite as each day completes —
        # so even a run that stalls partway has already saved its progress.
        chunk_result,fail_reasons=compute_daily_revenue_batch(
            client.base_url,sid,todo,client.xsrf,client.verify_ssl,
            request_gap_seconds=float(pace),on_progress=_on_progress)
        prog.empty(); status.empty()

        n_new_ok=sum(1 for d in todo if d in chunk_result)
        if not fail_reasons:
            st.success(f"✅ Computed {n_new_ok} new day(s) this run.")
        else:
            st.info(f"Computed {n_new_ok} new day(s); {len(fail_reasons)} failed this run.")
        if fail_reasons:
            with st.expander(f"⚠️ {len(fail_reasons)}/{len(todo)} day(s) failed — reasons"):
                for d in sorted(fail_reasons):
                    st.write(f"**{d}**: {fail_reasons[d]}")
            if len(fail_reasons)==len(todo):
                st.warning(
                    "Every day in this run failed — almost certainly still "
                    "rate-limited from a previous run. Try again later with a "
                    "longer pace, or wait a while before retrying."
                )
        # Re-read from SQLite so the table below reflects what this run just wrote.
        cached_days=_cached_revenue_days(elapsed_days)

    if cached_days:
        df_dr=pd.DataFrame([
            {"MonthNum":d.month,"kwh":cached_days[str(d)]["kwh"],
             "revenue_eur":cached_days[str(d)]["revenue_eur"]}
            for d in elapsed_days if str(d) in cached_days])
        mo_rev=df_dr.groupby("MonthNum").agg(
            Revenue_EUR=("revenue_eur","sum"),
            Energy_15min_kWh=("kwh","sum"),
            Days_Computed=("kwh","count")).reset_index()
        elapsed_by_month=(pd.Series([d.month for d in elapsed_days], name="MonthNum")
                          .value_counts().rename_axis("MonthNum")
                          .reset_index(name="Days_Elapsed"))
        merged=merged.merge(mo_rev,on="MonthNum",how="left")
        merged=merged.merge(elapsed_by_month,on="MonthNum",how="left")
        merged["CapturePrice"]=(merged["Revenue_EUR"]
                                / merged["Energy_15min_kWh"] * 1000)
        # A month with e.g. 1/24 days computed shows a real but tiny partial
        # sum in Revenue (€) — without this it reads as a complete monthly
        # total and looks like a bug when spot-checked against Intraday.
        merged["Coverage"]=merged.apply(
            lambda r: f"{int(r['Days_Computed'])}/{int(r['Days_Elapsed'])}"
            if pd.notna(r.get("Days_Computed")) and pd.notna(r.get("Days_Elapsed"))
            and r["Days_Elapsed"]>0 else "—", axis=1)
        n_partial=int(((merged["Days_Computed"]<merged["Days_Elapsed"])
                       & merged["Days_Computed"].notna()).sum())
        st.caption(f"✅ Revenue computed for {len(cached_days)} day(s) total, cached. "
                  + (f"⚠️ {n_partial} month(s) below only show a **partial** sum "
                     "— check the Coverage column." if n_partial else ""))
    else:
        merged["Revenue_EUR"]=pd.NA
        merged["CapturePrice"]=pd.NA

    # Chart
    fig=go.Figure()
    fig.add_bar(x=merged["Month"],y=merged["Budget_kWh"],
                name="Budget",marker_color="#334155")
    fig.add_bar(x=merged["Month"],y=merged["Energy_kWh"],
                name="Actual",marker_color="#f0b429")
    if frames:
        dc=pd.concat(frames).sort_values("dt").reset_index(drop=True)
        dc["Roll12"]=dc["kWh"].rolling(12,min_periods=3).mean()
        dy=dc[dc["yr"]==year_input]
        if not dy.empty:
            fig.add_scatter(x=[MONTH_LABELS[m-1] for m in dy["m"]],
                            y=dy["Roll12"].values,
                            mode="lines+markers",name="12-mo Rolling",
                            line=dict(color="#f472b6",width=2.5,dash="dot"))
    _base_layout(fig,f"{year_input} Monthly Energy vs Budget",
                 "Month","kWh",barmode="group")
    if year_input == today.year:
        st.caption(
            f"📅 **{MONTH_LABELS[today.month-1]} budget prorated** to day "
            f"{today.day} of {calendar.monthrange(today.year,today.month)[1]} "
            f"({today.day/calendar.monthrange(today.year,today.month)[1]:.0%} of month)."
        )
    st.plotly_chart(fig,use_container_width=True)

    # Delta chart
    dc2=[("🟢" if v>=0 else "🔴","#4ade80" if v>=0 else "#ff5f5f")
         for v in merged["Delta_kWh"].fillna(0)]
    fig2=go.Figure(go.Bar(x=merged["Month"],y=merged["Delta_kWh"],
        marker_color=[c[1] for c in dc2],name="Δ vs Budget"))
    _base_layout(fig2,"Monthly Delta (Actual − Budget)","Month","kWh")
    st.plotly_chart(fig2,use_container_width=True)

    # Summary table
    has_rev=merged["Revenue_EUR"].notna().any()
    dcols=["Month","Budget_kWh","Energy_kWh","Delta_kWh","Achievement_%"]
    fmt={"Budget_kWh":"{:,.0f}","Energy_kWh":"{:,.0f}",
         "Delta_kWh":"{:+,.0f}","Achievement_%":"{:.1f}%"}
    if has_rev:
        rename_map={"Revenue_EUR":"Revenue (€)",
                   "CapturePrice":"Capture Price (€/MWh)",
                   "Coverage":"Days Computed"}
        merged=merged.rename(columns=rename_map)
        dcols+=["Days Computed","Capture Price (€/MWh)","Revenue (€)"]
        fmt["Capture Price (€/MWh)"]="{:.2f}"
        fmt["Revenue (€)"]="{:,.0f}"
        ytd_rev=merged["Revenue (€)"].sum(skipna=True)
        st.metric("YTD Revenue",f"€ {ytd_rev:,.0f}",
                  help="Σ (15-min ENTSO-E price × generation) for months "
                       "you've computed. Months not yet computed show as "
                       "\"—\" — click 📊 Compute Revenue above. Check the "
                       "Days Computed column — a month showing e.g. 1/24 has "
                       "a real but tiny partial sum, not a wrong total.")

    st.subheader("Summary Table")
    st.dataframe(merged[dcols].style.format(fmt,na_rep="—"),
                 use_container_width=True)

    # PDF download
    conn=_get_db()
    alert_rows=conn.execute(
        "SELECT * FROM alerts ORDER BY ts DESC LIMIT 20").fetchall()
    conn.close()
    pvg=pvgis_df([year_input],ref_pr)
    merged2=merged.rename(columns={"Energy_kWh_x":"Energy_kWh"}) \
        if "Energy_kWh_x" in merged.columns else merged
    pdf_bytes=build_pdf_report(merged2, pvg, alert_rows, year_input)
    if pdf_bytes:
        st.download_button("📥 Download PDF Report",pdf_bytes,
            file_name=f"APM_Report_{year_input}.pdf",
            mime="application/pdf")

# ─────────────────────────────────────────────────────────────────────────────
# TAB: INTRADAY  (15-min + DAM price dual-axis)
# ─────────────────────────────────────────────────────────────────────────────
with t_intraday:
    st.header("15-min Production vs 15-min Day-Ahead Price")
    c1,c2 = st.columns([2,1])
    with c1:
        tgt_date=st.date_input("Day",value=date.today(),
                               key="id_date")
    with c2:
        show_rev=st.checkbox("Show revenue curve",value=True)

    _run_intraday = (st.button("🔄 Generate", key="btn_id")
                     or "intraday_date" not in st.session_state
                     or st.session_state.get("intraday_date") != str(tgt_date))

    if _run_intraday:
        st.session_state["intraday_date"] = str(tgt_date)

        # Prefer the local Excel exports over the live API: getKpiStation5min
        # doesn't exist for this account (HTTP 404, not just rate-limited),
        # so a live call can never do better than hourly-fallback resolution,
        # and getKpiStationHour is itself frequently rate-limited on top of
        # that. The local files are genuinely native 15-min per-inverter
        # readings — strictly better whenever they cover the requested day.
        jq = _local_prod_json_for_day(tgt_date)
        used_local = jq is not None

        client, stations = ensure_client()
        if not client or not stations:
            if not used_local:
                st.error("❌ FusionSolar connection failed — check secrets.toml.")
                st.stop()
            st.warning("⚠️ FusionSolar connection failed (event markers won't be shown) — "
                      "continuing with local Excel production data.")
        if used_local or (client and stations):
            sid = (stations[0].get("stationCode") or stations[0].get("plantCode")
                  if stations else None)
            if not used_local:
                with st.spinner("Fetching production data…"):
                    jq=api_15min(client.base_url,sid,tgt_date,client.xsrf,client.verify_ssl)
            with st.spinner("Fetching DAM prices…"):
                df_dam=dam_daily(tgt_date)

            raw=jq.get("data",[])
            src=jq.get("_src","hour")
            if raw and isinstance(raw,list) and "kpiList" in (raw[0] if isinstance(raw[0],dict) else {}):
                raw=raw[0]["kpiList"]
            df_p=_norm(raw)
            if df_p.empty:
                st.warning("No production data."); st.stop()

            tc=next((c for c in df_p.columns if "time" in c.lower() or "collect" in c.lower()),None)
            yc=next((c for c in df_p.columns if "inverterYield" in c or "activePower" in c
                     or "day_cap" in c or "power" in c.lower()),None)
            if not tc or not yc:
                st.error(f"Columns: {list(df_p.columns)}"); st.stop()

            df_p["dt"]=pd.to_datetime(df_p[tc],unit="ms",utc=True).dt.tz_convert(PLANT_TZ)
            df_p[yc]=pd.to_numeric(df_p[yc],errors="coerce")
            df_p=df_p.sort_values("dt").reset_index(drop=True)

            if len(df_p) >= 2:
                median_gap = df_p["dt"].diff().dropna().dt.total_seconds().median()
                interval_min = int(round(median_gap / 60))
            else:
                interval_min = 5

            if used_local:
                st.caption(f"📁 Using local Excel export (native {interval_min}-min, "
                          f"{len(df_p)} intervals) — 3 inverters summed.")
            elif src=="hour" and len(df_p)<=25:
                st.caption("⚠️ Live API data is hourly (5-min endpoint unavailable for this "
                          "account) — chart below is still shown at 15-min resolution by "
                          "repeating each hour's average across its four quarters.")
                interval_min = 60
            else:
                st.caption(f"✅ Native {interval_min}-min resolution ({len(df_p)} intervals) — bucketed to 15-min for the chart below.")

            # Chart always renders at 15-min resolution regardless of native
            # cadence. Sub-15-min data (5-min) is averaged down into each
            # 15-min bucket. Hourly-fallback data is coarser than 15-min, so
            # instead of dumping a whole hour's energy into one bucket (which
            # would spike that bucket to 4x true power and leave the other
            # three empty), each hourly reading is expanded into its four
            # constituent 15-min slots at the SAME kW value — the best
            # available estimate for each quarter given only an hourly
            # average — before bucketing.
            if interval_min > 15:
                df_p_exp = pd.concat([
                    df_p.assign(dt=df_p["dt"] + pd.Timedelta(minutes=m))
                    for m in range(0, interval_min, 15)
                ], ignore_index=True)
            else:
                df_p_exp = df_p
            df_p_exp["_bucket"] = df_p_exp["dt"].dt.floor("15min")
            bucket_kw = (df_p_exp.groupby("_bucket")[yc].mean().reset_index()
                        .rename(columns={"_bucket": "dt", yc: "kw_avg"}))

            dam_ok = not df_dam.empty and "price" in df_dam.columns

            if dam_ok and show_rev:
                # The core figures (kwh/revenue_eur/avg_price + the 15-min
                # bucket breakdown) come from the SAME function the Monthly
                # tab's batch compute uses — this tab must never maintain its
                # own independent copy of that formula again.
                r = _compute_day_revenue_from_frames(jq, df_dam)
                if r:
                    dr       = r["buckets"]   # columns: dt, kwh, price, rev
                    tot_kwh  = r["kwh"]
                    tot_rev  = r["revenue_eur"]
                    avg_p    = r["avg_price"]
                    pk_p     = df_dam["price"].max()
                    hi_thr   = df_dam["price"].quantile(0.75)

                    # Capture price/rate and top-quartile % are Intraday-only
                    # presentation extras (full-resolution, not bucketed) —
                    # not part of the stored figure, so fine to compute
                    # separately from the shared calc above.
                    # Normalize to UTC before merging — see comment in
                    # _compute_day_revenue_from_frames for why (dtype
                    # mismatch between differently-constructed tz-aware
                    # columns that pandas' merge_asof rejects outright).
                    df_dam_s = (df_dam[["dt","price"]]
                               .assign(dt=df_dam["dt"].dt.tz_convert("UTC"))
                               .sort_values("dt"))
                    left_p = (df_p[["dt",yc]]
                             .assign(dt=df_p["dt"].dt.tz_convert("UTC"))
                             .sort_values("dt"))
                    dr_full = pd.merge_asof(left_p, df_dam_s, on="dt",
                                            direction="backward",
                                            tolerance=pd.Timedelta("16min"))
                    cap_price = ((dr_full[yc] * dr_full["price"]).sum() /
                                 dr_full[yc].sum()) if dr_full[yc].sum() > 0 else np.nan
                    cap_rate  = cap_price / avg_p if avg_p > 0 else np.nan
                    hi_pct    = (dr_full.loc[dr_full["price"] >= hi_thr, yc].sum()
                                 / dr_full[yc].sum() * 100
                                 if dr_full[yc].sum() > 0 else 0)

                    m1,m2,m3,m4 = st.columns(4)
                    m1.metric("Total Production", f"{tot_kwh:,.0f} kWh")
                    m2.metric("Est. Revenue",      f"€ {tot_rev:,.1f}")
                    m3.metric("Avg DAM Price",     f"{avg_p:.1f} €/MWh")
                    m4.metric("Peak DAM Price",    f"{pk_p:.1f} €/MWh")
                    st.caption(f"📊 {hi_pct:.1f}% of production in top-quartile price "
                               f"window (≥{hi_thr:.1f} €/MWh).")
                    st.info(f"💡 **Capture Price:** {cap_price:.2f} €/MWh  |  "
                            f"**Capture Rate:** {cap_rate:.1%}  "
                            f"(1.0 = perfectly aligned with market)")

                    # Store the exact shared-calc result (including buckets)
                    # so the save button below persists identically to how
                    # compute_daily_revenue_batch would for this same day.
                    st.session_state["intraday_result"] = dict(
                        day=str(tgt_date), kwh=tot_kwh, revenue_eur=tot_rev,
                        avg_price=avg_p, buckets=dr, interval_min=interval_min)
                else:
                    st.info("Could not compute revenue for this day — "
                           "production/price data didn't line up.")

            fig=make_subplots(specs=[[{"secondary_y":True}]])
            fig.add_trace(go.Scatter(
                x=bucket_kw["dt"], y=bucket_kw["kw_avg"], fill="tozeroy",
                line=dict(color="#f0b429", width=1.5, shape="hv"),
                fillcolor="rgba(240,180,41,0.12)",
                name="Production (kW, 15-min)"),
                secondary_y=False)

            if dam_ok:
                fig.add_trace(go.Scatter(
                    x=df_dam["dt"], y=df_dam["price"],
                    mode="lines", line=dict(color="#3ecfcf", width=2, shape="hv"),
                    name="SMP €/MWh"),
                    secondary_y=True)
                if show_rev and "dr" in locals():
                    fig.add_trace(go.Bar(
                        x=dr["dt"], y=dr["rev"],
                        marker_color="rgba(167,139,250,0.5)",
                        name="Revenue (€/15-min)"),
                        secondary_y=True)
            else:
                st.info("DAM prices unavailable — production curve only.")

            events = pd.DataFrame()
            if client and stations:
                with st.spinner("Fetching FusionSolar events…"):
                    ev_raw = api_alarms_range(client.base_url, sid, tgt_date, tgt_date,
                                              client.xsrf, client.verify_ssl)
                    events = normalize_alarms(ev_raw)
                if not events.empty:
                    _add_event_markers(fig, events)

            _dual_layout(fig,
                f"Production (15-min kW) & DAM Price (15-min) — {tgt_date}",
                f"⚡ Production (kW)",
                "💰 Price (€/MWh)  ·  Revenue (€/15-min)")

            # Anchor both y-axes at 0 explicitly — rangemode="tozero" (set in
            # _dual_layout) only guarantees zero is *included*, not that it's
            # the axis edge, so on the secondary axis in particular the price
            # line could otherwise float above a non-zero bottom.
            power_max = bucket_kw["kw_avg"].max() if not bucket_kw.empty else 0
            fig.update_yaxes(range=[0, power_max * 1.1 if power_max > 0 else 1],
                             secondary_y=False)
            if dam_ok:
                right_series = [df_dam["price"]]
                if show_rev and "dr" in locals():
                    right_series.append(dr["rev"])
                right_all = pd.concat(right_series)
                right_lo = min(0, right_all.min())
                right_hi = right_all.max() * 1.1 if right_all.max() > 0 else 1
                fig.update_yaxes(range=[right_lo, right_hi], secondary_y=True)

            st.plotly_chart(fig, use_container_width=True)
            if not events.empty:
                st.caption(f"⚠️ {len(events)} FusionSolar event(s) on {tgt_date} — "
                          "hover the diamond markers near the baseline for details.")

            with st.expander("📋 15-min Revenue Breakdown"):
                if dam_ok and show_rev and "dr" in locals():
                    disp = dr[["dt","kwh","price","rev"]].copy()
                    disp["Time"] = disp["dt"].dt.strftime("%H:%M")
                    st.dataframe(
                        disp[["Time","kwh","price","rev"]].rename(columns={
                            "kwh":    "Energy (kWh)",
                            "price":  "SMP (€/MWh)",
                            "rev":    "Revenue (€)"
                        }).style.format({
                            "Energy (kWh)":  "{:.3f}",
                            "SMP (€/MWh)":   "{:.2f}",
                            "Revenue (€)":   "{:.4f}"
                        }),
                        use_container_width=True, hide_index=True)

    # ── Store daily revenue ───────────────────────────────────────────────────
    st.divider()
    res = st.session_state.get("intraday_result")
    if res:
        st.caption(f"Last calculated: **{res['day']}** — "
                   f"{res['kwh']:,.0f} kWh · €{res['revenue_eur']:,.2f}")
        if st.button("💾 Save daily revenue to log", key="btn_extract"):
            # Same persistence path compute_daily_revenue_batch uses — writes
            # daily_revenue AND revenue_15min, so a day saved from Intraday
            # is indistinguishable from one computed via the Monthly tab.
            _persist_day_revenue(date.fromisoformat(res["day"]), {
                "kwh": res["kwh"], "revenue_eur": res["revenue_eur"],
                "avg_price": res["avg_price"], "buckets": res["buckets"]})
            st.success(f"✅ Saved {res['day']} to revenue log.")

    # ── Revenue calendar ──────────────────────────────────────────────────────
    st.divider()
    st.subheader("📅 Daily Revenue Calendar")
    today = date.today()
    cal_months = [(today.year, today.month)]
    prev_m = today.month - 1 or 12
    prev_y = today.year if today.month > 1 else today.year - 1
    cal_months = [(prev_y, prev_m), (today.year, today.month)]

    conn = _get_db()
    rev_rows = conn.execute(
        "SELECT day, kwh, revenue_eur, avg_price FROM daily_revenue ORDER BY day"
    ).fetchall()
    conn.close()

    rev_map = {r[0]: {"kwh": r[1], "rev": r[2], "price": r[3]} for r in rev_rows}

    for cal_yr, cal_mo in cal_months:
        st.markdown(f"**{calendar.month_name[cal_mo]} {cal_yr}**")
        first_wd, n_days = calendar.monthrange(cal_yr, cal_mo)
        # Header
        cols = st.columns(7)
        for i, dn in enumerate(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]):
            cols[i].markdown(f"<center><small>{dn}</small></center>",
                             unsafe_allow_html=True)
        # Blanks before first day
        day_num = 1
        cells = [""] * first_wd
        while day_num <= n_days:
            d_str = f"{cal_yr}-{cal_mo:02d}-{day_num:02d}"
            r = rev_map.get(d_str)
            if r:
                cells.append(f"**{day_num}**\n\n€{r['rev']:.0f}")
            else:
                cells.append(str(day_num))
            day_num += 1
        # Pad to full weeks
        while len(cells) % 7:
            cells.append("")
        # Render rows
        for row_start in range(0, len(cells), 7):
            row_cells = cells[row_start:row_start+7]
            cols = st.columns(7)
            for i, cell in enumerate(row_cells):
                if cell:
                    cols[i].markdown(cell)
        st.caption(f"Monthly total: **€{sum(v['rev'] for k,v in rev_map.items() if k.startswith(f'{cal_yr}-{cal_mo:02d}')):.0f}**")
        st.divider()


# ─────────────────────────────────────────────────────────────────────────────
# TAB: HISTORICAL  (multi-year, PR, WCPR, degradation, irradiance)
# ─────────────────────────────────────────────────────────────────────────────
with t_hist:
    st.header("Historical Performance Deep-Dive")
    all_years=list(range(max(PLANT_START_YEAR, 2025), date.today().year+1))
    hist_years=st.multiselect("Years",options=all_years,default=all_years)

    if not hist_years:
        st.info("Select years above.")
    elif st.button("Load Historical Data",key="btn_hist"):
        client, stations = ensure_client()
        if not client or not stations:
            st.error("❌ FusionSolar connection failed — check secrets.toml.")
            st.stop()
        if client and stations:
            sid=stations[0].get("stationCode") or stations[0].get("plantCode")
            frames=[]
            sorted_years=sorted(hist_years)
            prog=st.progress(0,"Fetching…")
            done=0
            with concurrent.futures.ThreadPoolExecutor(max_workers=min(4,len(sorted_years))) as ex:
                fut_map={ex.submit(api_monthly,client.base_url,sid,yr,
                                   client.xsrf,client.verify_ssl):yr for yr in sorted_years}
                for fut in concurrent.futures.as_completed(fut_map):
                    yr=fut_map[fut]
                    df_y=fut.result()
                    if not df_y.empty:
                        df_y=df_y.copy(); df_y["_yr"]=yr; frames.append(df_y)
                    done+=1
                    prog.progress(done/len(sorted_years),f"Fetched {yr}")
            frames.sort(key=lambda f: f["_yr"].iloc[0])
            prog.empty()
            if not frames: st.warning("No data."); st.stop()

            df_all=pd.concat(frames,ignore_index=True)
            tc,ec=_resolve(df_all)
            if not tc or not ec: st.error(f"Cols: {list(df_all.columns)}"); st.stop()

            temp_col=next((c for c in df_all.columns if "temperature" in c.lower()),None)
            df_all["dt"]=pd.to_datetime(df_all[tc],unit="ms",utc=True).dt.tz_convert(PLANT_TZ)
            df_all["YM"]=df_all["dt"].dt.to_period("M").astype(str)
            df_all["m"]=df_all["dt"].dt.month
            df_all["yr"]=df_all["_yr"]
            df_all["kWh"]=pd.to_numeric(df_all[ec],errors="coerce")
            df_all=df_all.sort_values("dt").reset_index(drop=True)
            if temp_col:
                df_all["Temp"]=pd.to_numeric(df_all[temp_col],errors="coerce")

            pvg=pvgis_df(hist_years,ref_pr)
            df_mo=(df_all.groupby("YM")["kWh"].sum()
                   .reset_index().sort_values("YM").reset_index(drop=True))
            df_mo["m"]=df_mo["YM"].apply(lambda x:int(x.split("-")[1]))
            df_mo["yr"]=df_mo["YM"].apply(lambda x:int(x.split("-")[0]))
            df_mo["Year"]=df_mo["yr"]
            df_mo.rename(columns={"kWh":"Energy_kWh"},inplace=True)

            df_b=df_mo.merge(pvg[["YearMonth","GTI","T_amb","Expected_kWh"]].rename(
                columns={"YearMonth":"YM"}),on="YM",how="left")
            df_b["PR"]=df_b["Energy_kWh"]/(df_b["GTI"]*PLANT_PEAK_KW)
            df_b["WCPR"]=df_b.apply(lambda r:wcpr(r["Energy_kWh"],r["GTI"],r["T_amb"]),axis=1)
            df_b["Ratio_%"] = (
                df_b["Energy_kWh"].astype(float)
                / df_b["Expected_kWh"].astype(float).where(df_b["Expected_kWh"] != 0)
                * 100
            ).round(1)
            df_b["Roll12"]=df_b["Energy_kWh"].rolling(12,min_periods=3).mean()
            df_b["PR_R12"]=df_b["PR"].rolling(12,min_periods=3).mean()
            df_b["WCPR_R12"]=df_b["WCPR"].rolling(12,min_periods=3).mean()

            # Run alert checks
            run_alert_checks(df_b)

            # ① YoY chart
            st.subheader("① Year-over-Year Monthly Energy")
            fig1=go.Figure()
            for i,yr in enumerate(sorted(hist_years)):
                dy=df_mo[df_mo["yr"]==yr].sort_values("m")
                fig1.add_scatter(x=dy["m"],y=dy["Energy_kWh"],mode="lines+markers",
                    name=str(yr),line=dict(color=PALETTE[i%len(PALETTE)],width=2))
            bdf_r=get_budget(sorted(hist_years)[-1])
            fig1.add_scatter(x=list(range(1,13)),y=bdf_r["Budget_kWh"],
                mode="lines",name="Budget",line=dict(color="#475569",dash="dash"))
            fig1.add_scatter(x=df_b["m"],y=df_b["Roll12"],mode="lines",
                name="12-mo Rolling",line=dict(color="#f472b6",width=2,dash="dot"))
            _base_layout(fig1,"Monthly Energy per Year","Month","kWh")
            fig1.update_xaxes(tickmode="array",tickvals=list(range(1,13)),
                              ticktext=MONTH_LABELS)
            st.plotly_chart(fig1,use_container_width=True)

            # ② Cumulative
            st.subheader("② Cumulative Energy Since COD")
            df_b["Cum_MWh"]=df_b["Energy_kWh"].cumsum()/1000
            fig2=go.Figure(go.Scatter(x=df_b["YM"],y=df_b["Cum_MWh"],
                fill="tozeroy",line_color="#3ecfcf",
                fillcolor="rgba(62,207,207,0.12)",name="Cumulative MWh"))
            _base_layout(fig2,"Cumulative Energy (MWh)","","MWh")
            st.plotly_chart(fig2,use_container_width=True)
            st.metric("Total Production to Date",
                      f"{df_b['Cum_MWh'].iloc[-1]:,.1f} MWh")

            # ③ PR + WCPR trend
            st.subheader("③ Performance Ratio (PR) & Weather-Corrected PR (WCPR)")
            fig3=go.Figure()
            fig3.add_scatter(x=df_b["YM"],y=df_b["PR"],mode="lines+markers",
                line=dict(color="#60a5fa",width=1.5),name="Monthly PR",marker=dict(size=4))
            fig3.add_scatter(x=df_b["YM"],y=df_b["WCPR"],mode="lines+markers",
                line=dict(color="#f0b429",width=1.5,dash="dot"),
                name="Monthly WCPR",marker=dict(size=4))
            fig3.add_scatter(x=df_b["YM"],y=df_b["PR_R12"],mode="lines",
                name="PR 12-mo avg",line=dict(color="#a78bfa",width=2,dash="dot"))
            fig3.add_scatter(x=df_b["YM"],y=df_b["WCPR_R12"],mode="lines",
                name="WCPR 12-mo avg",line=dict(color="#fb923c",width=2,dash="dot"))
            fig3.add_hline(y=0.75,line_dash="dash",line_color="#4ade80",
                           annotation_text="Target 0.75")
            fig3.add_hline(y=0.65,line_dash="dot",line_color="#ff5f5f",
                           annotation_text="Alert 0.65")
            _base_layout(fig3,"PR & WCPR Trend","","PR / WCPR")
            fig3.update_yaxes(tickformat=".0%")
            st.plotly_chart(fig3,use_container_width=True)

            # ④ Irradiance benchmark
            st.subheader("④ Irradiance Benchmark — Actual vs PVGIS Expected")
            fig4=make_subplots(rows=2,cols=1,shared_xaxes=True,
                vertical_spacing=0.1,
                subplot_titles=["kWh: Actual vs Expected","Actual / Expected %"])
            fig4.add_bar(x=df_b["YM"],y=df_b["Expected_kWh"],
                name="PVGIS Expected",marker_color="#334155",row=1,col=1)
            fig4.add_bar(x=df_b["YM"],y=df_b["Energy_kWh"],
                name="Actual",marker_color="#f0b429",row=1,col=1)
            fig4.add_scatter(x=df_b["YM"],y=df_b["Roll12"],mode="lines",
                name="12-mo Actual",line=dict(color="#f472b6",width=2,dash="dot"),row=1,col=1)
            rc=["#4ade80" if v>=100 else "#fb923c" if v>=85 else "#ff5f5f"
                for v in df_b["Ratio_%"].fillna(0)]
            fig4.add_bar(x=df_b["YM"],y=df_b["Ratio_%"],
                name="Actual/Expected %",marker_color=rc,row=2,col=1)
            fig4.add_hline(y=100,line_dash="dash",line_color="#4ade80",row=2,col=1)
            fig4.add_hline(y=85,line_dash="dot",line_color="#ff5f5f",
                           annotation_text="85% alert",row=2,col=1)
            fig4.update_layout(barmode="group",height=500,
                paper_bgcolor=_BG,plot_bgcolor=_BG,font=dict(color=_TXT),
                legend=dict(bgcolor="rgba(0,0,0,0)"),margin=dict(t=55,b=40,l=10,r=10))
            fig4.update_xaxes(gridcolor=_GRD); fig4.update_yaxes(gridcolor=_GRD)
            st.plotly_chart(fig4,use_container_width=True)

            # ⑤ Degradation tracker
            st.subheader("⑤ Degradation Rate Analysis")
            deg=fit_degradation(df_mo)
            if deg:
                c1,c2,c3=st.columns(3)
                pct=deg["pct_per_year"]
                c1.metric("Degradation Rate",f"{pct:.2f}%/year",
                          delta="⚠️ Above typical" if abs(pct)>0.8 else "✅ Within range",
                          delta_color="inverse" if abs(pct)>0.8 else "normal",
                          help="Typical crystalline-Si: 0.5–0.7%/year")
                c2.metric("R²",f"{deg['r2']:.3f}")
                projected_20yr=(deg["intercept"]+deg["slope"]*
                                (sorted(hist_years)[-1]+20))
                c3.metric("Projected Yield in 20yr",
                          f"{projected_20yr:,.0f} kWh/kWp")

                da=deg["data"]
                x_fit=np.linspace(da["Year"].min(),da["Year"].max()+1,50)
                y_fit=deg["intercept"]+deg["slope"]*x_fit
                figD=go.Figure()
                figD.add_scatter(x=da["Year"],y=da["Specific_Yield"],
                    mode="markers+lines",name="Annual Specific Yield",
                    line=dict(color="#f0b429",width=2))
                figD.add_scatter(x=x_fit,y=y_fit,mode="lines",
                    name=f"Trend ({pct:.2f}%/yr)",
                    line=dict(color="#ff5f5f",dash="dash",width=2))
                # 90% CI band
                se=np.std(da["Specific_Yield"].values-
                          (deg["intercept"]+deg["slope"]*da["Year"].values))
                figD.add_scatter(x=np.concatenate([x_fit,x_fit[::-1]]),
                    y=np.concatenate([y_fit+1.645*se,(y_fit-1.645*se)[::-1]]),
                    fill="toself",fillcolor="rgba(255,95,95,0.08)",
                    line=dict(color="rgba(0,0,0,0)"),name="90% CI",showlegend=True)
                _base_layout(figD,"Annual Specific Yield & Degradation Trend",
                             "Year","kWh/kWp")
                st.plotly_chart(figD,use_container_width=True)
            else:
                st.info("Need ≥2 years of data for degradation analysis.")

            # ⑥ ML Anomaly detection
            st.subheader("⑥ Anomaly Detection (Isolation Forest)")
            df_an=run_anomaly_detection(df_all)
            if not df_an.empty:
                n_anom=(df_an["anomaly"]==-1).sum()
                st.metric("Anomalous months detected",str(n_anom))
                figA=go.Figure()
                normal=df_an[df_an["anomaly"]==1]
                anom  =df_an[df_an["anomaly"]==-1]
                figA.add_scatter(x=normal.index,y=normal["score"],
                    mode="markers",name="Normal",
                    marker=dict(color="#4ade80",size=7))
                figA.add_scatter(x=anom.index,y=anom["score"],
                    mode="markers",name="Anomaly",
                    marker=dict(color="#ff5f5f",size=10,symbol="x"))
                _base_layout(figA,"Anomaly Score by Month (lower = more anomalous)",
                             "Month index","Score")
                st.plotly_chart(figA,use_container_width=True)
                if n_anom>0:
                    add_alert("Warning","Anomaly",
                        f"Isolation Forest flagged {n_anom} anomalous months")
            else:
                st.info("Insufficient telemetry columns for ML anomaly detection. "
                        "Need efficiency, power_factor, or temperature data.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB: LOSS CASCADE  (IEC 61724 energy waterfall)
# ─────────────────────────────────────────────────────────────────────────────
with t_loss:
    st.header("Loss Cascade — IEC 61724 Energy Waterfall")
    st.caption("Decompose losses from theoretical yield to AC output. "
               "Configure soiling and wiring loss estimates below.")

    lc1,lc2,lc3=st.columns(3)
    soiling_pct = lc1.number_input("Soiling loss (%)",value=2.5,min_value=0.0,
                                    max_value=20.0,step=0.1,
                                    help="Annual average soiling loss estimate")/100
    wiring_pct  = lc2.number_input("DC wiring loss (%)",value=1.5,min_value=0.0,
                                    max_value=10.0,step=0.1)/100
    inv_eff     = lc3.number_input("Inverter efficiency (%)",value=97.5,
                                    min_value=80.0,max_value=100.0,step=0.1)/100

    cascade_year=st.selectbox("Year",options=list(range(PLANT_START_YEAR,
                                                        date.today().year+1)),
                               index=0,key="lc_year")

    if st.button("Compute Loss Cascade",key="btn_loss"):
        client, stations = ensure_client()
        if not client or not stations:
            st.error("❌ FusionSolar connection failed — check secrets.toml.")
            st.stop()
        if client and stations:
            sid=stations[0].get("stationCode") or stations[0].get("plantCode")
            with st.spinner("Fetching…"):
                df_yr=api_monthly(client.base_url,sid,cascade_year,
                                  client.xsrf,client.verify_ssl)
            if df_yr.empty: st.warning("No data."); st.stop()
            tc,ec=_resolve(df_yr)
            if not tc or not ec: st.error("Column detection failed."); st.stop()

            df_yr["dt"]=pd.to_datetime(df_yr[tc],unit="ms",utc=True).dt.tz_convert(PLANT_TZ)
            df_yr["m"]=df_yr["dt"].dt.month
            df_yr["kWh"]=pd.to_numeric(df_yr[ec],errors="coerce")
            actual_kwh=df_yr["kWh"].sum(skipna=True)

            # Annual GTI and T_amb
            gti_ann  =sum(PVGIS_GTI[m] for m in range(1,13))
            tamb_ann =sum(T_AMB[m] for m in range(1,13))/12

            # Ref yield = GTI × capacity (no losses)
            ref_yield      = gti_ann * PLANT_PEAK_KW

            # Temperature loss
            t_cell_avg = tamb_ann + (NOCT-20)/800 * (gti_ann*1000/8760)
            temp_loss_frac = abs(GAMMA * (t_cell_avg - 25))
            temp_loss_kwh  = ref_yield * temp_loss_frac

            # After temp
            after_temp = ref_yield - temp_loss_kwh

            # Soiling
            soil_loss  = after_temp * soiling_pct
            after_soil = after_temp - soil_loss

            # Inverter
            inv_loss   = after_soil * (1-inv_eff)
            after_inv  = after_soil * inv_eff

            # DC wiring
            wire_loss  = after_inv * wiring_pct
            after_wire = after_inv - wire_loss

            # Unexplained loss (residual between modelled and actual)
            unexplained = after_wire - actual_kwh
            unexplained = max(unexplained, 0)

            measures=["relative","relative","relative","relative","relative","total"]
            vals=[ref_yield,
                  -temp_loss_kwh, -soil_loss, -inv_loss, -wire_loss,
                  actual_kwh]
            labels=["Reference Yield",
                    f"Temperature Loss ({temp_loss_frac:.1%})",
                    f"Soiling Loss ({soiling_pct:.1%})",
                    f"Inverter Loss ({(1-inv_eff):.1%})",
                    f"DC Wiring Loss ({wiring_pct:.1%})",
                    "AC Output (Actual)"]
            colors_wf=(["#60a5fa"]+["#ff5f5f"]*4+["#4ade80"])

            fig_wf=go.Figure(go.Waterfall(
                orientation="v",measure=measures,
                x=labels,y=vals,
                connector=dict(line=dict(color="#334155",width=1)),
                increasing=dict(marker_color="#4ade80"),
                decreasing=dict(marker_color="#ff5f5f"),
                totals=dict(marker_color="#f0b429"),
                text=[f"{v:,.0f}" for v in vals],
                textposition="outside",
            ))
            _base_layout(fig_wf,f"IEC 61724 Loss Cascade — {cascade_year}","","kWh")
            fig_wf.update_layout(height=480)
            st.plotly_chart(fig_wf,use_container_width=True)

            pr_eff=actual_kwh/ref_yield if ref_yield>0 else 0
            c1,c2,c3,c4=st.columns(4)
            c1.metric("Reference Yield",f"{ref_yield:,.0f} kWh")
            c2.metric("AC Output",f"{actual_kwh:,.0f} kWh")
            c3.metric("System PR",f"{pr_eff:.3f}")
            c4.metric("Unexplained Residual",
                      f"{unexplained:,.0f} kWh",
                      help="Gap between bottom-up model and actual — "
                           "investigate shading, clipping, or curtailment.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB: FINANCIAL  (Revenue waterfall, Capture Rate, DSCR)
# ─────────────────────────────────────────────────────────────────────────────
with t_financial:
    st.header("Financial Performance & Covenant Monitor")

    if st.button("Load Financial Data",key="btn_fin"):
        client, stations = ensure_client()
        if not client or not stations:
            st.error("❌ FusionSolar connection failed — check secrets.toml.")
            st.stop()
        if client and stations:
            sid=stations[0].get("stationCode") or stations[0].get("plantCode")
            today=date.today()
            cur_yr  = today.year
            prev_yr = cur_yr - 1
            fin_years = [prev_yr, cur_yr]

            frames=[]
            with st.spinner("Fetching last year and current year…"):
                yr_data=api_monthly_years(client.base_url,sid,fin_years,
                                          client.xsrf,client.verify_ssl)
                for yr in fin_years:
                    df_y=yr_data[yr]
                    if not df_y.empty:
                        df_y=df_y.copy(); df_y["_yr"]=yr; frames.append(df_y)
            if not frames: st.warning("No data."); st.stop()

            df_all=pd.concat(frames,ignore_index=True)
            tc,ec=_resolve(df_all)
            df_all["dt"]=pd.to_datetime(df_all[tc],unit="ms",utc=True).dt.tz_convert(PLANT_TZ)
            df_all["m"]=df_all["dt"].dt.month; df_all["yr"]=df_all["_yr"]
            df_all["kWh"]=pd.to_numeric(df_all[ec],errors="coerce")
            df_mo=(df_all.groupby(["yr","m"])["kWh"].sum()
                   .reset_index().sort_values(["yr","m"]))
            df_mo["YM"]=df_mo.apply(lambda r:f"{int(r['yr'])}-{int(r['m']):02d}",axis=1)

            with st.spinner("Fetching monthly DAM prices…"):
                pairs=sorted({(int(row["yr"]),int(row["m"])) for _,row in df_mo.iterrows()
                             if date(int(row["yr"]),int(row["m"]),1)<today})
                prices=dam_monthly_avg_batch(pairs)
            df_mo["DAM"]=df_mo.apply(lambda r:prices.get((int(r["yr"]),int(r["m"]))),axis=1)
            df_mo["Revenue"]=df_mo.apply(
                lambda r:r["kWh"]*r["DAM"]/1000
                if pd.notna(r["DAM"]) and r["DAM"]>0 else pd.NA, axis=1)
            df_mo["OPEX_EUR"]=(fixed_opex/12 + df_mo["kWh"]/1000*var_opex)
            df_mo["EBITDA"]=df_mo["Revenue"].fillna(0)-df_mo["OPEX_EUR"]
            df_mo["Rev12"]  =df_mo["Revenue"].fillna(0).rolling(12,min_periods=6).sum()
            df_mo["OPEX12"] =df_mo["OPEX_EUR"].rolling(12,min_periods=6).sum()
            df_mo["DS12"]   =annual_debt
            df_mo["DSCR"]   =(df_mo["Rev12"]-df_mo["OPEX12"])/df_mo["DS12"]

            # ── YTD / MTD summary strip ──────────────────────────────────────
            st.subheader("① YTD & MTD Summary")
            cur_data = df_mo[df_mo["yr"]==cur_yr]
            prev_data= df_mo[df_mo["yr"]==prev_yr]

            ytd_rev  = cur_data["Revenue"].sum(skipna=True)
            ytd_kwh  = cur_data["kWh"].sum(skipna=True)
            ytd_rev_py=prev_data["Revenue"].sum(skipna=True)
            ytd_kwh_py=prev_data["kWh"].sum(skipna=True)

            mtd_row  = cur_data[cur_data["m"]==today.month]
            mtd_rev  = float(mtd_row["Revenue"].iloc[0]) if not mtd_row.empty and pd.notna(mtd_row["Revenue"].iloc[0]) else 0.0
            mtd_kwh  = float(mtd_row["kWh"].iloc[0])    if not mtd_row.empty else 0.0
            mtd_price= float(mtd_row["DAM"].iloc[0])    if not mtd_row.empty and pd.notna(mtd_row["DAM"].iloc[0]) else 0.0

            # Prorated MTD budget
            bdf_cur = get_budget(cur_yr)
            days_gone = today.day
            days_total= calendar.monthrange(today.year, today.month)[1]
            mtd_budg_kwh = float(bdf_cur.loc[bdf_cur["MonthNum"]==today.month,"Budget_kWh"].iloc[0])
            mtd_budg_pro = mtd_budg_kwh * days_gone / days_total

            fa,fb,fc_,fd = st.columns(4)
            fa.metric(f"YTD Revenue ({cur_yr})", f"€ {ytd_rev:,.0f}",
                      delta=f"€ {ytd_rev-ytd_rev_py:+,.0f} vs {prev_yr}" if ytd_rev_py else None)
            fb.metric(f"YTD Production ({cur_yr})", f"{ytd_kwh/1e3:,.1f} MWh",
                      delta=f"{(ytd_kwh-ytd_kwh_py)/1e3:+,.1f} MWh vs {prev_yr}" if ytd_kwh_py else None)
            fc_.metric(f"MTD Revenue ({MONTH_LABELS[today.month-1]})",
                       f"€ {mtd_rev:,.0f}", delta=f"Avg SMP: {mtd_price:.1f} €/MWh")
            fd.metric(f"MTD Production vs Pro-rated Budget",
                      f"{mtd_kwh:,.0f} kWh",
                      delta=f"{mtd_kwh-mtd_budg_pro:+,.0f} kWh vs budget",
                      delta_color="normal" if mtd_kwh>=mtd_budg_pro else "inverse")

            st.divider()
            st.subheader("② Annual Revenue Waterfall")
            wa_year=st.selectbox("Waterfall year",options=fin_years,
                                  index=len(fin_years)-1,key="wa_yr")
            dy=df_mo[df_mo["yr"]==wa_year]

            pvg_yr=pvgis_df([wa_year],ref_pr)
            exp_rev_base=(pvg_yr["Expected_kWh"].sum()
                          *prices.get((wa_year,6),80)/1000
                          if prices.get((wa_year,6)) else None)

            act_rev =dy["Revenue"].sum(skipna=True)
            act_kwh =dy["kWh"].sum(skipna=True)
            exp_kwh =pvg_yr["Expected_kWh"].sum()
            avg_dam =dy["DAM"].mean(skipna=True)

            if avg_dam and not np.isnan(avg_dam):
                irr_shortfall =(exp_kwh-act_kwh)*avg_dam/1000
                irr_shortfall = max(0, irr_shortfall)
                wf_measures=["absolute","relative","relative","total"]
                wf_x=["PVGIS Expected Revenue",
                       "Irradiance / PR Shortfall",
                       "Price Variance",
                       "Actual Revenue"]
                wf_y=[exp_kwh*avg_dam/1000,
                      -irr_shortfall,
                      0,
                      act_rev]
                fig_rev=go.Figure(go.Waterfall(
                    orientation="v",measure=wf_measures,
                    x=wf_x,y=wf_y,
                    increasing=dict(marker_color="#4ade80"),
                    decreasing=dict(marker_color="#ff5f5f"),
                    totals=dict(marker_color="#f0b429"),
                    text=[f"€{v:,.0f}" for v in wf_y],
                    textposition="outside"))
                _base_layout(fig_rev,f"Revenue Waterfall — {wa_year}","","€")
                st.plotly_chart(fig_rev,use_container_width=True)

            # ── Capture Rate trend ──
            st.subheader("③ Monthly Revenue & EBITDA")
            df_mo["Capture_Price"]=df_mo["DAM"]  # proxy — improve with 15-min data
            df_mo["Capture_Rate"]=df_mo["Capture_Price"]/df_mo["DAM"]
            figC=go.Figure()
            figC.add_scatter(x=df_mo["YM"],y=df_mo["Revenue"].fillna(0),
                mode="lines+markers",name="Monthly Revenue (€)",
                line=dict(color="#f0b429",width=2))
            figC.add_scatter(x=df_mo["YM"],y=df_mo["EBITDA"],
                mode="lines",name="EBITDA (€)",
                line=dict(color="#4ade80",width=1.5,dash="dot"))
            _base_layout(figC,"Monthly Revenue & EBITDA","","€")
            st.plotly_chart(figC,use_container_width=True)

            # ── DSCR Chart ──
            st.subheader("④ Rolling 12-Month DSCR")
            df_dscr=df_mo.dropna(subset=["DSCR"])
            figDSCR=go.Figure()
            figDSCR.add_scatter(x=df_dscr["YM"],y=df_dscr["DSCR"],
                mode="lines+markers",name="DSCR (trailing 12mo)",
                line=dict(color="#60a5fa",width=2),marker=dict(size=5))
            figDSCR.add_hline(y=1.20,line_dash="dash",line_color="#fb923c",
                              annotation_text="1.20x Lock-up trigger")
            figDSCR.add_hline(y=1.10,line_dash="dot",line_color="#ff5f5f",
                              annotation_text="1.10x Default trigger")
            figDSCR.add_hrect(y0=0,y1=1.10,fillcolor="rgba(255,95,95,0.05)",
                              line_width=0)
            _base_layout(figDSCR,"Debt Service Coverage Ratio (DSCR)","","DSCR")
            st.plotly_chart(figDSCR,use_container_width=True)

            last_dscr=df_dscr["DSCR"].iloc[-1] if not df_dscr.empty else np.nan
            if not np.isnan(last_dscr):
                col_a,col_b=st.columns(2)
                col_a.metric("Latest DSCR (trailing 12mo)",f"{last_dscr:.2f}x",
                    delta="⚠️ Below lock-up" if last_dscr<1.20 else "✅ Above lock-up",
                    delta_color="inverse" if last_dscr<1.20 else "normal")
                # Stress test
                rev_needed_for_lockup=(1.20*annual_debt+df_mo["OPEX12"].iloc[-1])
                rev_gap=rev_needed_for_lockup-df_mo["Rev12"].iloc[-1]
                col_b.metric("Revenue gap to lock-up trigger",
                    f"€ {abs(rev_gap):,.0f}",
                    delta="Surplus" if rev_gap<0 else "Deficit",
                    delta_color="normal" if rev_gap<0 else "inverse")

            if last_dscr<1.20 and not np.isnan(last_dscr):
                add_alert("Critical","DSCR",
                    f"Trailing 12-month DSCR = {last_dscr:.2f}x — below 1.20x lock-up threshold")

# ─────────────────────────────────────────────────────────────────────────────
# TAB: FORECAST  (Open-Meteo 7-day production + revenue)
# ─────────────────────────────────────────────────────────────────────────────
with t_forecast:
    st.header("7-Day Production & Revenue Forecast")
    st.caption("Production forecast: Open-Meteo GHI → GTI (×1.15 proxy) → yield model. "
               "Revenue forecast uses next-day ENTSO-E Day-Ahead price if available, else trailing avg.")

    if st.button("Load Forecast",key="btn_fc"):
        with st.spinner("Fetching Open-Meteo forecast…"):
            df_fc=fetch_forecast()
        if df_fc.empty:
            err = st.session_state.get("_forecast_error","")
            st.warning(
                "⚠️ Could not fetch forecast from Open-Meteo.\n\n"
                + (f"**Error:** `{err}`\n\n" if err else "")
                + "**Possible fixes:**\n"
                "- Add `[proxy] https = \"http://proxy-rwe-de.energy.local:8080\"` to secrets.toml\n"
                "- Check that `api.open-meteo.com` is reachable from your network\n"
                "- Try: `curl https://api.open-meteo.com/v1/forecast` in terminal"
            )
            st.stop()

        # Aggregate to daily
        df_fc["Date"]=df_fc["dt"].dt.date
        df_day=(df_fc.groupby("Date")
                .agg(Yield_kWh=("Yield_kWh","sum"),
                     Peak_GHI=("GHI_Wm2","max"),
                     Avg_Temp=("T_amb","mean"))
                .reset_index())
        df_day["Date"]=pd.to_datetime(df_day["Date"])

        # Fetch next-day DAM (D+1 published by ADMIE ~13:00)
        fc_dam={}
        with st.spinner("Checking ENTSO-E for next-day prices…"):
            fc_dates=df_day["Date"].dt.date.tolist()
            dam_results=dam_daily_batch(fc_dates)
            for d in fc_dates:
                p=dam_results.get(d)
                if p is not None and not p.empty and "price" in p.columns:
                    fc_dam[d]=p["price"].mean()
        if fc_dam:
            df_day["DAM"]=df_day["Date"].dt.date.map(fc_dam)
        else:
            # Fall back to trailing monthly average
            trail=dam_monthly_avg(date.today().year, date.today().month)
            df_day["DAM"]=trail

        df_day["Rev_EUR"]=df_day.apply(
            lambda r: r["Yield_kWh"]*r["DAM"]/1000
            if pd.notna(r["DAM"]) and r["DAM"]>0 else pd.NA, axis=1)

        # Chart
        figF=make_subplots(specs=[[{"secondary_y":True}]])
        figF.add_bar(x=df_day["Date"],y=df_day["Yield_kWh"],
            name="Forecast Yield (kWh)",marker_color="#f0b429",
            secondary_y=False)
        if df_day["Rev_EUR"].notna().any():
            figF.add_scatter(x=df_day["Date"],y=df_day["Rev_EUR"],
                mode="lines+markers",name="Est. Revenue (€)",
                line=dict(color="#3ecfcf",width=2),secondary_y=True)
        _dual_layout(figF,"7-Day Production & Revenue Forecast",
            "⚡ Yield (kWh)","💰 Est. Revenue (€)")
        st.plotly_chart(figF,use_container_width=True)

        # Hourly profile chart
        st.subheader("Hourly Yield Profile (7 days)")
        figH=go.Figure()
        for i,d in enumerate(df_fc["dt"].dt.date.unique()[:7]):
            dh=df_fc[df_fc["dt"].dt.date==d]
            figH.add_scatter(x=dh["dt"].dt.hour,y=dh["Yield_kWh"],
                mode="lines",name=str(d),
                line=dict(color=PALETTE[i%len(PALETTE)],width=1.5))
        _base_layout(figH,"Hourly Yield Profile by Day","Hour of Day","kWh")
        st.plotly_chart(figH,use_container_width=True)

        # Summary table
        st.dataframe(df_day.rename(columns={
            "Date":"Date","Yield_kWh":"Forecast kWh",
            "Peak_GHI":"Peak GHI (W/m²)","Avg_Temp":"Avg Temp (°C)",
            "DAM":"DAM (€/MWh)","Rev_EUR":"Est. Revenue (€)"})
            .style.format({"Forecast kWh":"{:,.0f}","Peak GHI (W/m²)":"{:.0f}",
                           "Avg Temp (°C)":"{:.1f}","DAM (€/MWh)":"{:.2f}",
                           "Est. Revenue (€)":"{:,.0f}"},na_rep="—"),
            use_container_width=True,hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
# TAB: HEALTH  (Live diagnostics + strings + availability + soiling)
# ─────────────────────────────────────────────────────────────────────────────
with t_health:
    st.header("Equipment Health & Diagnostics")

    h_tab1, h_tab2, h_tab3 = st.tabs(["🔌 Live Diagnostics",
                                       "📅 Availability Log",
                                       "🌿 Soiling Optimiser"])

    with h_tab1:
        if st.button("Run Live Diagnostics",key="btn_diag"):
            client, stations = ensure_client()
            if not client:
                st.error("❌ Cannot connect to FusionSolar — check credentials in secrets.toml.")
                st.stop()
            if not stations:
                st.error("❌ No stations returned from FusionSolar API.")
                st.stop()
            if client and stations:
                sid=stations[0].get("stationCode") or stations[0].get("plantCode")
                with st.spinner("Fetching device list…"):
                    j_devs=_post(client.s,
                                 f"{client.base_url}/thirdData/getDevList",
                                 {"stationCodes":sid})
                inv_ids=[d["id"] for d in j_devs.get("data",[])
                         if d.get("devTypeId") in INV_DEV_TYPE_IDS]
                if not inv_ids:
                    st.warning("No inverters found.")
                else:
                    df_kpi=api_realtime(client.base_url,1,tuple(inv_ids[:3]),
                                        client.xsrf,client.verify_ssl)
                    if df_kpi.empty:
                        st.warning("No real-time KPI data.")
                    else:
                        tech_map={"dataItemMap.active_power":"AC Power (kW)",
                                  "dataItemMap.temperature":"Temp (°C)",
                                  "dataItemMap.elec_freq":"Freq (Hz)",
                                  "dataItemMap.power_factor":"PF",
                                  "dataItemMap.efficiency":"Eff %"}
                        df_disp=df_kpi.rename(columns=tech_map)
                        avail=[v for v in tech_map.values() if v in df_disp.columns]
                        st.subheader("Inverter Telemetry")
                        st.dataframe(df_disp[avail],use_container_width=True)
                        # Persist snapshot for failure trend analysis
                        _save_telemetry_snapshot(df_kpi)

                        # Temp alert
                        if "Temp (°C)" in df_disp.columns:
                            max_temp=pd.to_numeric(df_disp["Temp (°C)"],errors="coerce").max()
                            if max_temp>INV_TEMP_WARN:
                                add_alert("Warning","Temperature",
                                    f"Inverter temperature {max_temp:.1f}°C > "
                                    f"{INV_TEMP_WARN}°C threshold")
                                st.warning(f"⚠️ Inverter temperature {max_temp:.1f}°C "
                                           f"exceeds {INV_TEMP_WARN}°C threshold!")

                        st.divider()
                        st.subheader("DC String Analysis")
                        string_cols=[c for c in df_kpi.columns if "pv" in c and "_i" in c]
                        if not string_cols:
                            st.info("No DC string columns found (pattern: pv*_i).")
                        else:
                            s_vals=df_kpi[string_cols].iloc[0].astype(float).sort_values(ascending=False)
                            active=s_vals[s_vals>0.1]
                            mean_a=active.mean() if len(active)>0 else 0
                            bar_c=["#ff5f5f" if v<=0.1
                                   else "#f0b429" if v<mean_a*0.88
                                   else "#4ade80" for v in s_vals.values]
                            figS=go.Figure(go.Bar(x=s_vals.index,y=s_vals.values,
                                                  marker_color=bar_c))
                            if mean_a>0:
                                figS.add_hline(y=mean_a,line_dash="dash",
                                    line_color="#60a5fa",
                                    annotation_text=f"Mean: {mean_a:.2f}A")
                            _base_layout(figS,"DC String Currents","String","Amperes")
                            st.plotly_chart(figS,use_container_width=True)

                            if len(active)>1:
                                cv=active.std()/active.mean()
                                low=active[active<active.mean()*0.88]
                                c1,c2,c3=st.columns(3)
                                c1.metric("Active Strings",len(active))
                                c2.metric("Mean Current",f"{active.mean():.2f} A")
                                c3.metric("CV (variation)",f"{cv:.1%}")
                                if cv>STRING_CV_WARN:
                                    add_alert("Warning","Strings",
                                        f"String CV {cv:.1%} > {STRING_CV_WARN:.0%} — "
                                        f"{len(low)} strings underperforming")
                                    st.error(f"⚠️ High variation (CV {cv:.1%}) — "
                                             f"{len(low)} string(s) >12% below mean: "
                                             f"{list(low.index)}")
                                else:
                                    st.success(f"✅ Strings balanced (CV {cv:.1%}).")

    with h_tab2:
        st.subheader("Downtime Log & Availability")
        with st.expander("➕ Log Downtime Event"):
            dn_c1,dn_c2=st.columns(2)
            dn_start=dn_c1.date_input("Start date",key="dn_start")
            dn_end  =dn_c2.date_input("End date",  key="dn_end")
            dn_inv  =st.text_input("Inverter / component","INV-01",key="dn_inv")
            dn_kwh  =st.number_input("Estimated lost kWh",0.0,step=10.0,key="dn_kwh")
            dn_cause=st.text_input("Cause","Grid fault",key="dn_cause")
            if st.button("Save Downtime",key="dn_save"):
                conn=_get_db()
                conn.execute("INSERT INTO downtime VALUES(NULL,?,?,?,?,?)",
                    (dn_start.isoformat(),dn_end.isoformat(),
                     dn_inv,dn_kwh,dn_cause))
                conn.commit(); conn.close()
                st.success("Logged.")

        conn=_get_db()
        rows=conn.execute("SELECT * FROM downtime ORDER BY start_dt DESC").fetchall()
        conn.close()
        if rows:
            df_dn=pd.DataFrame(rows,columns=["id","Start","End","Inverter",
                                              "Lost kWh","Cause"])
            st.dataframe(df_dn.drop(columns="id"),use_container_width=True,
                         hide_index=True)
            total_lost=df_dn["Lost kWh"].sum()
            st.metric("Total Logged Lost Production",f"{total_lost:,.0f} kWh")
        else:
            st.info("No downtime events logged yet.")

    with h_tab3:
        st.subheader("Soiling Loss & Cleaning Optimisation")
        st.caption("Model the financial break-even point for a cleaning campaign.")

        so_c1,so_c2,so_c3=st.columns(3)
        clean_cost =so_c1.number_input("Cleaning cost (€)",value=500.0,step=50.0)
        yield_rec  =so_c2.number_input("Yield recovery (%)",value=2.5,step=0.1)/100
        dam_ref    =so_c3.number_input("DAM reference price (€/MWh)",value=80.0,step=1.0)

        # Revenue recovered per cleaning
        daily_kwh_est=PLANT_PEAK_KW*4.5  # rough 4.5 peak-sun-hours
        monthly_kwh  =daily_kwh_est*30
        rec_kwh_mo   =monthly_kwh*yield_rec
        rec_rev_mo   =rec_kwh_mo*dam_ref/1000

        # Payback in months
        payback_mo = clean_cost/rec_rev_mo if rec_rev_mo>0 else float("inf")

        c1,c2,c3=st.columns(3)
        c1.metric("Monthly recovered energy",f"{rec_kwh_mo:,.0f} kWh")
        c2.metric("Monthly recovered revenue",f"€ {rec_rev_mo:,.0f}")
        c3.metric("Payback period",
                  f"{payback_mo:.1f} months" if payback_mo<99 else "Not viable",
                  delta="✅ Viable" if payback_mo<3 else "⚠️ Check frequency")

        # Soiling accumulation curve
        days=np.arange(0,180)
        # Exponential soiling model: loss = A(1 - exp(-k·d))
        A_soil=0.05; k_soil=0.02
        soil_curve=A_soil*(1-np.exp(-k_soil*days))
        rev_loss_curve=soil_curve*monthly_kwh/30*dam_ref/1000

        figSo=go.Figure()
        figSo.add_scatter(x=days,y=soil_curve*100,name="Soiling loss (%)",
            line=dict(color="#f0b429",width=2))
        figSo.add_scatter(x=days,y=rev_loss_curve,name="Daily Revenue Loss (€)",
            line=dict(color="#ff5f5f",width=2),yaxis="y2")
        figSo.add_vline(x=payback_mo*30 if payback_mo<6 else 90,
                        line_dash="dash",line_color="#4ade80",
                        annotation_text="Recommended cleaning trigger")
        figSo.update_layout(paper_bgcolor=_BG,plot_bgcolor=_BG,
            font=dict(color=_TXT),hovermode="x unified",
            yaxis=dict(title="Soiling Loss (%)",gridcolor=_GRD,
                       title_font=dict(color="#f0b429"),tickfont=dict(color="#f0b429")),
            yaxis2=dict(title="Daily Revenue Loss (€)",overlaying="y",side="right",
                        gridcolor="rgba(0,0,0,0)",
                        title_font=dict(color="#ff5f5f"),tickfont=dict(color="#ff5f5f")),
            title=dict(text="Soiling Accumulation & Revenue Impact",
                       font=dict(color=_TXT,size=15)),
            legend=dict(bgcolor="rgba(0,0,0,0)"),margin=dict(t=55,b=40,l=70,r=70))
        figSo.update_xaxes(title_text="Days Since Last Cleaning",gridcolor=_GRD)
        st.plotly_chart(figSo,use_container_width=True)

        # Log cleaning event
        with st.expander("➕ Log Cleaning Event"):
            cl_dt=st.date_input("Cleaning date",key="cl_dt")
            cl_cost=st.number_input("Cost (€)",value=clean_cost,step=50.0,key="cl_cost")
            cl_rec=st.number_input("Measured yield recovery (%)",value=2.5,
                                    step=0.1,key="cl_rec")
            cl_notes=st.text_input("Notes","Full array cleaning",key="cl_notes")
            if st.button("Save Cleaning Event",key="cl_save"):
                conn=_get_db()
                conn.execute("INSERT INTO cleaning VALUES(NULL,?,?,?,?)",
                    (cl_dt.isoformat(),cl_cost,cl_rec,cl_notes))
                conn.commit(); conn.close()
                st.success("Logged.")

        conn=_get_db()
        cl_rows=conn.execute("SELECT * FROM cleaning ORDER BY dt DESC").fetchall()
        conn.close()
        if cl_rows:
            df_cl=pd.DataFrame(cl_rows,columns=["id","Date","Cost (€)",
                                                  "Yield Recovery (%)","Notes"])
            st.dataframe(df_cl.drop(columns="id"),use_container_width=True,
                         hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
# TAB: OPEX  (cost tracker + OPEX/MWh trend)
# ─────────────────────────────────────────────────────────────────────────────
with t_opex:
    st.header("OPEX Tracker")
    st.caption("Log O&M costs and track cost-per-MWh over time.")

    OPEX_CATS=["Routine O&M","Corrective Maintenance","Insurance",
               "Land Lease","Cleaning","Other"]

    with st.expander("➕ Add OPEX Entry"):
        op_c1,op_c2,op_c3=st.columns(3)
        op_dt  =op_c1.date_input("Date",key="op_dt")
        op_cat =op_c2.selectbox("Category",OPEX_CATS,key="op_cat")
        op_amt =op_c3.number_input("Amount (€)",0.0,step=100.0,key="op_amt")
        op_desc=st.text_input("Description","",key="op_desc")
        if st.button("Save OPEX",key="op_save"):
            conn=_get_db()
            conn.execute("INSERT INTO opex VALUES(NULL,?,?,?,?)",
                (op_dt.isoformat(),op_cat,op_desc,op_amt))
            conn.commit(); conn.close()
            st.success("OPEX entry saved.")

    conn=_get_db()
    opex_rows=conn.execute(
        "SELECT * FROM opex ORDER BY dt DESC").fetchall()
    conn.close()

    if opex_rows:
        df_op=pd.DataFrame(opex_rows,
                           columns=["id","Date","Category","Description","Amount (€)"])
        df_op["Date"]=pd.to_datetime(df_op["Date"])

        # Summary metrics
        total_opex=df_op["Amount (€)"].sum()
        ytd_opex=df_op[df_op["Date"].dt.year==date.today().year]["Amount (€)"].sum()
        c1,c2=st.columns(2)
        c1.metric("Total OPEX Logged",f"€ {total_opex:,.0f}")
        c2.metric("YTD OPEX",f"€ {ytd_opex:,.0f}")

        # Monthly trend
        df_op["YM"]=df_op["Date"].dt.to_period("M").astype(str)
        df_op_m=df_op.groupby(["YM","Category"])["Amount (€)"].sum().reset_index()
        figOP=go.Figure()
        for cat in OPEX_CATS:
            sub=df_op_m[df_op_m["Category"]==cat]
            if not sub.empty:
                figOP.add_bar(x=sub["YM"],y=sub["Amount (€)"],name=cat)
        _base_layout(figOP,"Monthly OPEX by Category","Month","€",barmode="stack")
        st.plotly_chart(figOP,use_container_width=True)

        # OPEX/MWh (need production data)
        if st.button("Calculate OPEX/MWh",key="opex_mwh"):
            client, stations = ensure_client()
            if not client or not stations:
                st.error("❌ FusionSolar connection failed — check secrets.toml credentials.")
                st.stop()
            if client and stations:
                sid=stations[0].get("stationCode") or stations[0].get("plantCode")
                frames=[]
                opex_years=list(range(PLANT_START_YEAR,date.today().year+1))
                with st.spinner(f"Fetching {len(opex_years)} years…"):
                    yr_data=api_monthly_years(client.base_url,sid,opex_years,
                                              client.xsrf,client.verify_ssl)
                for yr in opex_years:
                    df_y=yr_data[yr]
                    if not df_y.empty:
                        df_y=df_y.copy(); df_y["_yr"]=yr; frames.append(df_y)
                if frames:
                    df_all=pd.concat(frames,ignore_index=True)
                    tc,ec=_resolve(df_all)
                    if tc and ec:
                        df_all["dt"]=pd.to_datetime(df_all[tc],unit="ms",
                            utc=True).dt.tz_convert(PLANT_TZ)
                        df_all["YM"]=df_all["dt"].dt.to_period("M").astype(str)
                        df_all["kWh"]=pd.to_numeric(df_all[ec],errors="coerce")
                        df_prod_m=df_all.groupby("YM")["kWh"].sum().reset_index()
                        df_op_tot=df_op.groupby("YM")["Amount (€)"].sum().reset_index()
                        df_ratio=df_prod_m.merge(df_op_tot,on="YM",how="inner")
                        df_ratio["OPEX_MWh"]=(df_ratio["Amount (€)"]
                                              /df_ratio["kWh"]*1000)
                        figOM=go.Figure(go.Bar(x=df_ratio["YM"],
                            y=df_ratio["OPEX_MWh"],
                            marker_color=["#ff5f5f" if v>20 else "#4ade80"
                                          for v in df_ratio["OPEX_MWh"]]))
                        figOM.add_hline(y=15,line_dash="dash",line_color="#fb923c",
                            annotation_text="15 €/MWh guideline")
                        _base_layout(figOM,"OPEX per MWh","Month","€/MWh")
                        st.plotly_chart(figOM,use_container_width=True)

        st.subheader("OPEX Ledger")
        st.dataframe(df_op.drop(columns="id").style.format({"Amount (€)":"{:,.2f}"}),
                     use_container_width=True,hide_index=True)
    else:
        st.info("No OPEX entries yet. Add entries above to start tracking.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB: ALERTS  (alert feed + acknowledge / resolve)
# ─────────────────────────────────────────────────────────────────────────────
with t_alerts:
    st.header("Alert Feed")
    st.caption("Rule-based and ML-generated alerts. Acknowledge or resolve below.")

    conn=_get_db()
    all_alerts=conn.execute(
        "SELECT id,ts,severity,category,message,status FROM alerts "
        "ORDER BY ts DESC").fetchall()
    conn.close()

    if not all_alerts:
        st.success("✅ No alerts — all clear.")
    else:
        df_al=pd.DataFrame(all_alerts,
            columns=["id","Timestamp","Severity","Category","Message","Status"])

        # Metrics
        n_open =(df_al["Status"]=="Open").sum()
        n_crit =(df_al["Severity"]=="Critical").sum()
        n_warn =(df_al["Severity"]=="Warning").sum()
        c1,c2,c3=st.columns(3)
        c1.metric("Open Alerts",str(n_open))
        c2.metric("Critical",str(n_crit))
        c3.metric("Warnings",str(n_warn))

        # Filter
        filt=st.selectbox("Filter by status",["All","Open","Acknowledged","Resolved"],
                          key="al_filt")
        df_show=df_al if filt=="All" else df_al[df_al["Status"]==filt]

        # Colour severity
        def _sev_icon(s):
            return {"Critical":"🔴","Warning":"🟡","Info":"🔵"}.get(s,"⚪")
        df_show=df_show.copy()
        df_show["Sev"]=df_show["Severity"].apply(_sev_icon)+" "+df_show["Severity"]

        st.dataframe(df_show[["Timestamp","Sev","Category","Message","Status"]]
                     .rename(columns={"Sev":"Severity"}),
                     use_container_width=True,hide_index=True)

        # Acknowledge / Resolve
        st.divider()
        c1,c2=st.columns(2)
        alert_ids=[str(r[0]) for r in all_alerts if r[5]=="Open"]
        if alert_ids:
            sel_id=c1.selectbox("Select alert ID",alert_ids,key="al_sel")
            new_st=c2.selectbox("New status",["Acknowledged","Resolved"],key="al_st")
            if st.button("Update Alert",key="al_upd"):
                conn=_get_db()
                conn.execute("UPDATE alerts SET status=? WHERE id=?",
                             (new_st,int(sel_id)))
                conn.commit(); conn.close()
                st.success(f"Alert #{sel_id} → {new_st}")
                st.rerun()

        # Clear all resolved
        if st.button("🗑️ Clear all Resolved alerts",key="al_clr"):
            conn=_get_db()
            conn.execute("DELETE FROM alerts WHERE status='Resolved'")
            conn.commit(); conn.close()
            st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# TAB: FAILURE ANALYTICS  (historical telemetry trends → predictive signals)
# ─────────────────────────────────────────────────────────────────────────────
with t_failure:
    st.header("🩺 Predictive Failure Analytics")
    st.caption(
        "Historical trends across inverter temperature, power factor, efficiency, "
        "grid frequency and AC power — the early-warning signals for inverter and "
        "component failures. Data is built up each time you run **Live Diagnostics**."
    )

    try:
        conn = _get_db()
        telem_rows = conn.execute(
            "SELECT ts, inverter_id, signal, value FROM telemetry_history ORDER BY ts"
        ).fetchall()
        conn.close()
    except Exception as e:
        st.error(f"❌ Could not read telemetry database: `{e}`")
        telem_rows = []

    if not telem_rows:
        st.info(
            "No telemetry history yet. Go to **🛠️ Health → Live Diagnostics** and click "
            "**Run Live Diagnostics** — each run saves a snapshot. After a few runs "
            "(ideally over days/weeks) the trends here will be populated."
        )
        st.divider()
        st.subheader("What to look for — failure predictor guide")
        guide = [
            ("🌡️ Inverter Temperature rising", "Gradual °C rise above 60°C over weeks = fan failure, vent blockage, or internal component degradation. Sudden spike = cooling system fault. Lead time before failure: days to weeks."),
            ("⚡ Power Factor declining", "PF drop below 0.92 indicates grid interaction problems, capacitor bank degradation, or reactive power setpoint drift. Often precedes inverter trip events."),
            ("📉 Efficiency trend downward", "Efficiency dropping > 1% over 3 months (weather-corrected) suggests IGBT degradation, DC-side resistance increase (corroded connectors), or transformer issues."),
            ("🔌 Grid Frequency deviation", "Sustained deviation from 50 Hz (even 0.1 Hz) indicates grid instability that stresses inverter internals. Repeated excursions correlate with premature capacitor failure."),
            ("🔄 AC Power clipping", "Power plateau at rated capacity during low-irradiance periods = measurement anomaly or inverter de-rating. During high irradiance it's normal clipping."),
            ("📊 Multi-signal correlation", "The most reliable failure signal: temperature rising AND efficiency falling AND power factor dipping simultaneously = systemic degradation, not noise."),
        ]
        for title, detail in guide:
            with st.expander(title):
                st.write(detail)

    else:
        df_tel = pd.DataFrame(telem_rows, columns=["ts","inverter_id","signal","value"])
        df_tel["ts"] = pd.to_datetime(df_tel["ts"])
        df_tel["value"] = pd.to_numeric(df_tel["value"], errors="coerce")

        all_signals = sorted(df_tel["signal"].unique())
        all_invs    = sorted(df_tel["inverter_id"].unique())

        fa1, fa2 = st.columns([3, 1])
        with fa2:
            sel_inv = st.multiselect("Inverter(s)", all_invs,
                                     default=all_invs[:3] if len(all_invs)>=3 else all_invs,
                                     key="fa_inv")
        with fa1:
            date_range = st.date_input("Date range",
                value=(df_tel["ts"].dt.date.min(), df_tel["ts"].dt.date.max()),
                key="fa_dates")

        if not sel_inv:
            st.warning("Select at least one inverter.")
            st.stop()

        df_f = df_tel[df_tel["inverter_id"].isin(sel_inv)].copy()
        if len(date_range) == 2:
            df_f = df_f[(df_f["ts"].dt.date >= date_range[0]) &
                        (df_f["ts"].dt.date <= date_range[1])]

        if df_f.empty:
            st.warning("No data in selected range."); st.stop()

        # ── Summary health score per inverter ──
        st.subheader("① Inverter Health Score")
        st.caption("Composite score 0–100 based on how far each signal is from its warning threshold. "
                   "Below 70 = amber; below 50 = red.")

        def _health_score(df_inv: pd.DataFrame) -> float:
            scores = []
            for sig_col, meta in FAILURE_SIGNALS.items():
                sig_label = meta["label"]
                sub = df_inv[df_inv["signal"] == sig_label]
                if sub.empty or meta["warn"] is None: continue
                recent_val = sub.sort_values("ts")["value"].iloc[-1]
                low_bad = meta.get("low_bad", False)
                warn, crit = meta["warn"], meta["crit"]
                if low_bad:
                    if recent_val >= warn:   s = 100.0
                    elif recent_val >= crit: s = 50.0
                    else:                    s = 10.0
                else:
                    if recent_val <= warn:   s = 100.0
                    elif recent_val <= crit: s = 50.0
                    else:                   s = 10.0
                scores.append(s)
            return float(np.mean(scores)) if scores else 100.0

        score_cols = st.columns(min(len(sel_inv), 5))
        for i, inv in enumerate(sel_inv[:5]):
            df_i = df_f[df_f["inverter_id"] == inv]
            sc = _health_score(df_i)
            col = score_cols[i % len(score_cols)]
            delta_c = "normal" if sc >= 70 else "inverse"
            col.metric(f"INV {inv[-6:]}", f"{sc:.0f}/100",
                       delta="🟢 Good" if sc>=70 else ("🟡 Watch" if sc>=50 else "🔴 Critical"),
                       delta_color=delta_c)

        st.divider()

        # ── Per-signal historical trend charts ──
        st.subheader("② Signal Trend Charts")
        st.caption("Each panel shows raw readings + 7-day rolling average + warning/critical bands.")

        n_sig = len([s for s in FAILURE_SIGNALS.values() if
                     df_f["signal"].eq(s["label"]).any()])
        if n_sig == 0:
            st.info("No matching telemetry signals found in history.")
        else:
            for sig_col, meta in FAILURE_SIGNALS.items():
                sig_label = meta["label"]
                sub = df_f[df_f["signal"] == sig_label].copy()
                if sub.empty: continue

                st.markdown(f"**{sig_label}**")
                fig = go.Figure()

                # One trace per inverter
                for j, inv in enumerate(sel_inv):
                    dv = sub[sub["inverter_id"] == inv].sort_values("ts")
                    if dv.empty: continue
                    # Rolling mean (7-day window over irregular samples: use time-based)
                    dv = dv.set_index("ts")
                    dv["roll"] = dv["value"].rolling("7D", min_periods=1).mean()
                    dv = dv.reset_index()

                    fig.add_scatter(x=dv["ts"], y=dv["value"],
                        mode="markers", name=f"{inv} (raw)",
                        marker=dict(color=PALETTE[j % len(PALETTE)], size=5, opacity=0.5))
                    fig.add_scatter(x=dv["ts"], y=dv["roll"],
                        mode="lines", name=f"{inv} 7-day avg",
                        line=dict(color=PALETTE[j % len(PALETTE)], width=2))

                # Warning / critical bands
                warn_v = meta.get("warn")
                crit_v = meta.get("crit")
                low_bad = meta.get("low_bad", False)
                x_range = [df_f["ts"].min(), df_f["ts"].max()]
                if warn_v is not None:
                    fig.add_scatter(x=x_range, y=[warn_v, warn_v],
                        mode="lines", name="⚠️ Warning",
                        line=dict(color="#fb923c", dash="dash", width=1.5))
                if crit_v is not None:
                    fig.add_scatter(x=x_range, y=[crit_v, crit_v],
                        mode="lines", name="🔴 Critical",
                        line=dict(color="#ff5f5f", dash="dot", width=1.5))

                # Shaded danger zone
                if warn_v is not None and crit_v is not None:
                    if low_bad:
                        fig.add_hrect(y0=0, y1=crit_v,
                            fillcolor="rgba(255,95,95,0.06)", line_width=0)
                        fig.add_hrect(y0=crit_v, y1=warn_v,
                            fillcolor="rgba(251,146,60,0.04)", line_width=0)
                    else:
                        fig.add_hrect(y0=crit_v, y1=1e9,
                            fillcolor="rgba(255,95,95,0.06)", line_width=0)
                        fig.add_hrect(y0=warn_v, y1=crit_v,
                            fillcolor="rgba(251,146,60,0.04)", line_width=0)

                _base_layout(fig, "", "Date", f"{sig_label} ({meta['unit']})", height=260)
                fig.update_layout(margin=dict(t=10, b=30, l=10, r=10),
                                  legend=dict(orientation="h", y=1.02))
                st.plotly_chart(fig, use_container_width=True)

        st.divider()

        # ── Multi-signal correlation heatmap ──
        st.subheader("③ Cross-Signal Correlation Heatmap")
        st.caption("Correlation between telemetry signals across all inverters and time. "
                   "Strong negative correlation between temperature and efficiency is a key failure precursor.")

        pivot = df_f.pivot_table(index="ts", columns="signal", values="value",
                                 aggfunc="mean").resample("1D").mean().dropna(how="all")
        if len(pivot.columns) >= 2 and len(pivot) >= 3:
            corr = pivot.corr()
            fig_corr = go.Figure(go.Heatmap(
                z=corr.values,
                x=corr.columns.tolist(),
                y=corr.index.tolist(),
                colorscale="RdBu",
                zmid=0,
                zmin=-1, zmax=1,
                text=np.round(corr.values, 2).astype(str),
                texttemplate="%{text}",
                hovertemplate="%{y} vs %{x}: %{z:.3f}<extra></extra>",
            ))
            _base_layout(fig_corr, "Signal Correlation Matrix", height=380)
            fig_corr.update_layout(margin=dict(t=30, b=80, l=120, r=10))
            st.plotly_chart(fig_corr, use_container_width=True)

            # Highlight concerning correlations
            concerning = []
            for i in range(len(corr.columns)):
                for j in range(i+1, len(corr.columns)):
                    c = corr.iloc[i, j]
                    pair = f"{corr.columns[i]} ↔ {corr.index[j]}"
                    if c < -0.5:
                        concerning.append((pair, c, "🔴 Strong negative — investigate"))
                    elif c > 0.85 and i != j:
                        concerning.append((pair, c, "🟡 Redundant sensors or collinearity"))
            if concerning:
                st.markdown("**Notable correlations:**")
                for pair, val, note in concerning:
                    st.markdown(f"- `{pair}`: **{val:.2f}** — {note}")
        else:
            st.info("Need ≥2 signals and ≥3 days of data for correlation analysis.")

        st.divider()

        # ── Exceedance count over time ──
        st.subheader("④ Threshold Exceedance Log")
        st.caption("Count of readings that breached warning or critical thresholds per day.")

        exceedances = []
        for sig_col, meta in FAILURE_SIGNALS.items():
            sig_label = meta["label"]
            warn_v = meta.get("warn")
            crit_v = meta.get("crit")
            low_bad = meta.get("low_bad", False)
            sub = df_f[df_f["signal"] == sig_label].copy()
            if sub.empty or warn_v is None: continue
            sub["Date"] = sub["ts"].dt.date
            for _, row in sub.iterrows():
                v = row["value"]
                if low_bad:
                    level = ("Critical" if v < crit_v else
                             "Warning"  if v < warn_v else None)
                else:
                    level = ("Critical" if v > crit_v else
                             "Warning"  if v > warn_v else None)
                if level:
                    exceedances.append({
                        "Date": row["ts"],
                        "Inverter": row["inverter_id"],
                        "Signal": sig_label,
                        "Value": round(v, 3),
                        "Level": level,
                        "Threshold": warn_v if level=="Warning" else crit_v,
                    })

        if exceedances:
            df_exc = pd.DataFrame(exceedances).sort_values("Date", ascending=False)

            n_crit = (df_exc["Level"] == "Critical").sum()
            n_warn = (df_exc["Level"] == "Warning").sum()
            ec1, ec2 = st.columns(2)
            ec1.metric("Critical exceedances", str(n_crit))
            ec2.metric("Warning exceedances",  str(n_warn))

            # Time series of exceedance count
            df_exc["DateOnly"] = df_exc["Date"].dt.date
            daily_exc = df_exc.groupby(["DateOnly","Level"]).size().reset_index(name="Count")
            fig_exc = go.Figure()
            for level, col in [("Critical","#ff5f5f"),("Warning","#fb923c")]:
                sub = daily_exc[daily_exc["Level"]==level]
                if not sub.empty:
                    fig_exc.add_bar(x=sub["DateOnly"], y=sub["Count"],
                                    name=level, marker_color=col)
            _base_layout(fig_exc, "Daily Threshold Exceedances",
                         "Date", "Count", barmode="stack", height=240)
            st.plotly_chart(fig_exc, use_container_width=True)

            st.dataframe(
                df_exc[["Date","Inverter","Signal","Value","Threshold","Level"]]
                .style.apply(lambda row: [
                    "color:#ff5f5f" if row["Level"]=="Critical"
                    else "color:#fb923c" if row["Level"]=="Warning"
                    else "" for _ in row], axis=1),
                use_container_width=True, hide_index=True)

            # Auto-alert for critical exceedances
            if n_crit > 0:
                sigs = df_exc[df_exc["Level"]=="Critical"]["Signal"].unique()
                add_alert("Critical","FailureSignal",
                    f"{n_crit} critical threshold exceedance(s) detected in: "
                    f"{', '.join(sigs)}")
        else:
            st.success("✅ No threshold exceedances in selected range.")

        st.divider()

        # ── Failure probability estimate ──
        st.subheader("⑤ Failure Risk Estimate")
        st.caption("Simple heuristic model based on signal trend velocity. "
                   "Not a certified maintenance tool — use as a discussion aid.")

        risk_items = []
        for sig_col, meta in FAILURE_SIGNALS.items():
            sig_label = meta["label"]
            warn_v = meta.get("warn")
            if warn_v is None: continue
            sub = df_f[df_f["signal"] == sig_label].sort_values("ts")
            if len(sub) < 4: continue
            # Fit linear trend to last 30 readings
            recent = sub.tail(30)
            x_num = (recent["ts"] - recent["ts"].min()).dt.total_seconds().values
            y_val = recent["value"].values
            if np.std(x_num) == 0: continue
            try:
                from scipy.stats import linregress
                sl, ic, r2, _, _ = linregress(x_num, y_val)
            except Exception:
                continue
            # Days until threshold breach at current velocity
            low_bad = meta.get("low_bad", False)
            current = y_val[-1]
            crit_v = meta.get("crit", warn_v)
            if sl == 0: continue
            secs_to_breach = (crit_v - current) / sl if not low_bad else (current - crit_v) / (-sl)
            days_to_breach = secs_to_breach / 86400
            risk_items.append({
                "Signal": sig_label,
                "Current": round(current, 2),
                "Trend": f"{'↑' if sl>0 else '↓'} {abs(sl*86400):.3f}/day",
                "Days to Critical": round(days_to_breach, 0) if 0 < days_to_breach < 365 else None,
                "R²": round(r2 if hasattr(r2,'__float__') else r2**2, 3),
            })

        if risk_items:
            df_risk = pd.DataFrame(risk_items)
            df_risk["Risk"] = df_risk["Days to Critical"].apply(
                lambda d: "🔴 HIGH (<30d)" if d is not None and d < 30
                else ("🟡 MEDIUM (30–90d)" if d is not None and d < 90
                      else "🟢 Low / stable"))
            st.dataframe(
                df_risk.style.apply(lambda row: [
                    "color:#ff5f5f" if "HIGH" in str(row["Risk"])
                    else "color:#fb923c" if "MEDIUM" in str(row["Risk"])
                    else "color:#4ade80" for _ in row], axis=1),
                use_container_width=True, hide_index=True)

            high_risk = df_risk[df_risk["Risk"].str.contains("HIGH")]
            if not high_risk.empty:
                add_alert("Warning","FailureRisk",
                    f"Trend model predicts critical breach in <30 days for: "
                    f"{', '.join(high_risk['Signal'].tolist())}")
        else:
            st.info("Need ≥4 readings per signal to compute trend velocity.")

        st.divider()

        # ── Raw telemetry table + export ──
        with st.expander("📋 Raw Telemetry History"):
            st.dataframe(df_f[["ts","inverter_id","signal","value"]].rename(columns={
                "ts":"Timestamp","inverter_id":"Inverter",
                "signal":"Signal","value":"Value"}),
                use_container_width=True, hide_index=True)
            csv = df_f[["ts","inverter_id","signal","value"]].to_csv(index=False)
            st.download_button("⬇️ Export CSV", csv,
                file_name="telemetry_history.csv", mime="text/csv")

        if st.button("🗑️ Clear telemetry history", key="clr_tel"):
            conn = _get_db()
            conn.execute("DELETE FROM telemetry_history")
            conn.commit(); conn.close()
            st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# TAB: EVENTS  (FusionSolar device alarms — table, timeline, breakdown)
# ─────────────────────────────────────────────────────────────────────────────
with t_events:
    st.header("🚨 FusionSolar Events Analytics")
    st.caption(
        "Device alarms/events pulled live from FusionSolar (thirdData/getAlarmList). "
        "Field names can vary slightly by API version/tenant — use the raw response "
        "expander below to confirm the column mapping against your account."
    )

    ec1, ec2 = st.columns([2, 1])
    with ec1:
        ev_range = st.date_input("Date range",
            value=(date.today() - timedelta(days=30), date.today()),
            max_value=date.today(), key="ev_dates")
    with ec2:
        ev_sev = st.multiselect("Severity", ALARM_LEVEL_ORDER,
                                default=ALARM_LEVEL_ORDER, key="ev_sev")

    if not isinstance(ev_range, tuple) or len(ev_range) != 2:
        st.info("Pick a start and end date.")
        st.stop()
    ev_start, ev_end = ev_range

    _run_events = (st.button("🔄 Load Events", key="btn_events")
                  or "events_raw" not in st.session_state
                  or st.session_state.get("events_range") != (str(ev_start), str(ev_end)))

    if _run_events:
        client, stations = ensure_client()
        if not client or not stations:
            st.error("❌ FusionSolar connection failed — check secrets.toml.")
            st.stop()
        sid = stations[0].get("stationCode") or stations[0].get("plantCode")
        with st.spinner(f"Fetching events {ev_start} → {ev_end}…"):
            st.session_state["events_raw"] = api_alarms_range(
                client.base_url, sid, ev_start, ev_end,
                client.xsrf, client.verify_ssl)
        st.session_state["events_range"] = (str(ev_start), str(ev_end))

    ev_raw = st.session_state.get("events_raw", pd.DataFrame())
    events = normalize_alarms(ev_raw)

    if events.empty:
        st.info(
            "No events found for the selected range — or the FusionSolar alarm "
            "endpoint returned data in an unrecognised schema. Check the raw "
            "response expander below."
        )
    else:
        ev_f = events[events["Severity"].isin(ev_sev)] if ev_sev else events.iloc[0:0]

        st.subheader("① Summary")
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Total Events", str(len(ev_f)))
        k2.metric("Critical", str((ev_f["Severity"] == "Critical").sum()))
        k3.metric("Major", str((ev_f["Severity"] == "Major").sum()))
        latest = ev_f["dt"].max() if not ev_f.empty else None
        k4.metric("Most Recent",
                  latest.strftime("%Y-%m-%d %H:%M") if pd.notna(latest) else "—")

        st.divider()
        st.subheader("② Event Timeline")
        if ev_f.empty:
            st.info("No events match the selected severity filter.")
        else:
            figT = go.Figure()
            for sev in ALARM_LEVEL_ORDER:
                sub = ev_f[ev_f["Severity"] == sev]
                if sub.empty:
                    continue
                figT.add_scatter(
                    x=sub["dt"], y=[sev]*len(sub), mode="markers", name=sev,
                    marker=dict(size=10, symbol="diamond",
                               color=ALARM_LEVEL_COLOR.get(sev, "#94a3b8"),
                               line=dict(width=1, color="#0e1117")),
                    customdata=sub[["Device", "Alarm", "Cause"]].values,
                    hovertemplate="<b>%{customdata[1]}</b><br>%{customdata[0]}<br>"
                                 "%{customdata[2]}<br>%{x}<extra></extra>")
            _base_layout(figT, "Events over time", "Time", "Severity", showlegend=False)
            figT.update_yaxes(categoryorder="array",
                              categoryarray=list(reversed(ALARM_LEVEL_ORDER)))
            st.plotly_chart(figT, use_container_width=True)

        st.divider()
        st.subheader("③ Breakdown")
        b1, b2 = st.columns(2)
        with b1:
            sev_counts = ev_f["Severity"].value_counts().reindex(ALARM_LEVEL_ORDER).dropna()
            figS = go.Figure(go.Bar(
                x=sev_counts.index, y=sev_counts.values,
                marker_color=[ALARM_LEVEL_COLOR.get(s, "#94a3b8") for s in sev_counts.index]))
            _base_layout(figS, "By Severity", "Severity", "Count", showlegend=False)
            st.plotly_chart(figS, use_container_width=True)
        with b2:
            dev_counts = ev_f["Device"].value_counts().head(10)
            figD = go.Figure(go.Bar(x=dev_counts.values, y=dev_counts.index,
                                    orientation="h", marker_color="#3ecfcf"))
            _base_layout(figD, "Top Devices by Event Count", "Count", "", showlegend=False)
            st.plotly_chart(figD, use_container_width=True)

        st.divider()
        st.subheader("④ Event Log")
        tbl = ev_f.copy()
        tbl["Time"] = tbl["dt"].dt.strftime("%Y-%m-%d %H:%M")
        tbl = tbl[["Time", "Device", "Severity", "Alarm", "Cause", "Status"]] \
              .sort_values("Time", ascending=False)
        st.dataframe(tbl, use_container_width=True, hide_index=True)
        st.download_button("📥 Download events CSV", tbl.to_csv(index=False).encode(),
            file_name=f"fusionsolar_events_{ev_start}_{ev_end}.csv", mime="text/csv")

    with st.expander("🔍 Raw response (debug column mapping)"):
        if ev_raw is None or ev_raw.empty:
            st.caption("No raw rows fetched yet — click **Load Events** above.")
        else:
            st.caption(f"Resolved columns: `{_resolve_alarm_cols(ev_raw)}`")
            st.dataframe(ev_raw.head(20), use_container_width=True)

# ─────────────────────────────────────────────────────────────────────────────
# TAB: IPTO DIAGNOSTICS  (live connectivity probe + troubleshooting guide)
# ─────────────────────────────────────────────────────────────────────────────
with t_ipto:
    st.header("📡 ENTSO-E Market Data Diagnostics")
    st.caption(
        "Greek SMP prices powering the financial tabs are sourced from the "
        "ENTSO-E Transparency Platform (document A44, Day-Ahead, Greece bidding zone). "
        "Use this panel to verify connectivity and inspect the local price cache."
    )

    # Last error banner
    last_err = st.session_state.get("_entsoe_last_error")
    if last_err:
        st.error(f"**Last error recorded this session:**\n\n`{last_err}`")
    else:
        st.success("No errors recorded this session.")

    st.divider()

    # ── Live connectivity probe ───────────────────────────────────────────────
    st.subheader("Live Connectivity Probe")
    probe_col1, probe_col2 = st.columns([1, 2])
    probe_timeout = probe_col1.slider("Timeout (s)", 5, 30, 10, key="probe_timeout")

    if probe_col2.button("▶️ Run ENTSO-E Probe", key="btn_probe"):
        with st.spinner("Probing ENTSO-E API…"):
            result = probe_entsoe(timeout=probe_timeout)

        if result.get("user_message"):
            if result.get("price_parseable"):
                st.success(result["user_message"])
            elif result.get("reachable"):
                st.warning(result["user_message"])
            else:
                st.error(result["user_message"])

        detail_rows = [
            ("API reachable",      "✅ Yes" if result["reachable"] else "❌ No"),
            ("HTTP status",        str(result.get("status_code") or "N/A")),
            ("Error type",         result.get("error_type") or "None"),
            ("Prices parsed",      "✅ Yes" if result["price_parseable"] else "❌ No"),
            ("Periods returned",   str(result.get("n_periods", 0))),
            ("Sample avg price",   f"{result['sample_price']:.2f} €/MWh"
                                   if result.get("sample_price") else "N/A"),
            ("Detected NS",        result.get("detected_ns", "N/A")),
            ("Root tag",           result.get("root_tag", "N/A")),
            ("ACK reason code",    result.get("ack_code", "N/A")),
            ("ACK reason text",    result.get("ack_reason", "N/A")),
        ]
        st.dataframe(pd.DataFrame(detail_rows, columns=["Check", "Result"]),
                     use_container_width=True, hide_index=True)

        if result.get("raw_xml"):
            with st.expander("📄 Raw XML response (first 800 chars)"):
                st.code(result["raw_xml"], language="xml")

    st.divider()

    # ── Troubleshooting ───────────────────────────────────────────────────────
    st.subheader("Troubleshooting")

    with st.expander("❌ No token / 401 Unauthorised"):
        st.markdown("""
**Symptom:** `NO_TOKEN` or `INVALID_TOKEN` error type

**Fix:**
1. Register free at [transparency.entsoe.eu](https://transparency.entsoe.eu)
2. Go to **My Account Settings → Web API Security Token → Generate Token**
3. Add to `.streamlit/secrets.toml`:
```toml
[entsoe]
api_key = "your-uuid-token-here"
```
4. Restart the app — tokens look like `a1b2c3d4-e5f6-7890-abcd-ef1234567890`
""")

    with st.expander("⏱️ Timeout / DNS failure"):
        st.markdown("""
**Symptom:** `TIMEOUT` or `DNS_FAILURE` error type

**Fixes:**
1. Increase the timeout slider and re-run the probe.
2. Ensure `web-api.tp.entsoe.eu` (port 443) is reachable from your server.
3. If behind a corporate firewall, whitelist `tp.entsoe.eu`.
4. Test from terminal: `curl -I https://web-api.tp.entsoe.eu/api`
""")

    with st.expander("⚠️ Connected but 0 price points returned"):
        st.markdown("""
**Symptom:** API returns HTTP 200 but no prices parsed

**Possible causes:**
- The test date has no published prices yet (Day-Ahead prices published ~13:00 CET day before)
- The XML namespace changed — check [ENTSO-E API docs](https://transparency.entsoe.eu/content/static_content/Static%20content/web%20api/Guide.html)
- The Greece bidding zone EIC may have changed (currently `10YGR-HTSO-----Y`)
""")

    st.divider()

    # ── Local price cache ─────────────────────────────────────────────────────
    st.subheader("Local Price Cache (SQLite)")
    st.caption(
        "Monthly average prices are cached locally after each successful ENTSO-E fetch. "
        "Cached values are used as fallback if the API is temporarily unavailable."
    )

    conn = _get_db()
    cache_rows = conn.execute(
        "SELECT ym, avg_price_eur_mwh, fetched_ts FROM price_cache ORDER BY ym DESC"
    ).fetchall()
    conn.close()

    if cache_rows:
        df_cache = pd.DataFrame(cache_rows,
                                columns=["Month", "Avg Price (€/MWh)", "Fetched At"])
        st.dataframe(df_cache.style.format({"Avg Price (€/MWh)": "{:.2f}"}),
                     use_container_width=True, hide_index=True)
        st.metric("Cached months", len(cache_rows))

        if st.button("🗑️ Clear price cache", key="clr_cache"):
            conn = _get_db()
            conn.execute("DELETE FROM price_cache")
            conn.commit(); conn.close()
            st.rerun()
    else:
        st.info(
            "No prices cached yet. Prices are cached automatically as you use "
            "the Monthly and Financial tabs."
        )

    st.divider()

    with st.expander("📖 ENTSO-E API Reference"):
        st.markdown(f"""
**Base endpoint:** `https://web-api.tp.entsoe.eu/api`

**Key parameters for Greek SMP:**

| Parameter | Value |
|-----------|-------|
| `documentType` | `A44` (Price Document) |
| `processType` | `A01` (Day Ahead) |
| `in_Domain` | `{ENTSOE_ZONE}` |
| `out_Domain` | `{ENTSOE_ZONE}` |
| `periodStart` / `periodEnd` | UTC, format `YYYYMMDDHHII` |
| `securityToken` | your API token |

**Resolution:** 60-min (hourly) for Greece Day-Ahead prices

**Publication schedule:** Day-Ahead results published ~13:00 CET for the following day

**Docs:** [ENTSO-E Transparency Platform API Guide](https://transparency.entsoe.eu/content/static_content/Static%20content/web%20api/Guide.html)
""")
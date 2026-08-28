"""
Backfill daily_revenue/revenue_15min using LOCALLY EXPORTED inverter Excel
files instead of the FusionSolar API (which is still rate-limited,
failCode=407). The user downloaded per-inverter 15-min "Active power(kW)"
exports for com1-1/com1-2/com1-3 covering activation through today; this
script sums the 3 inverters per 15-min timestamp to get site-level power,
converts to energy (kWh = kW * 0.25h, matching the app's own convention),
and reuses the REAL apm_app.py revenue functions (AST-extracted, not
retyped) so results land in the exact same tables the app reads, computed
the exact same way as a normal API-sourced day.

Files use two different Start-Time conventions depending on season:
  - DST months: "YYYY-MM-DD HH:MM:SS DST"  (Athens EEST, UTC+3)
  - Standard months: "YYYY-MM-DD HH:MM:SS"  (Athens EET,  UTC+2, no suffix)
This is Huawei's own resolved local time, not re-derived — so the suffix
is used directly to pick the correct UTC offset (ambiguous=True/False in
tz_localize) instead of relying on Python's own DST inference.

These export files also don't declare a valid <dimension> in their sheet
XML (openpyxl read_only reports max_row=1), so max_row must be forced high
enough to actually read the data - iter_rows(max_row=200000) below.

Usage:
    python _local_prod_backfill.py --compare 2026-08-27   # validation only
    python _local_prod_backfill.py --run                  # backfill missing days
"""
import argparse
import ast
import glob
import os
import sqlite3
import tomllib
from datetime import date, datetime, timedelta, timezone
from typing import Optional

import openpyxl
import pandas as pd

DB_PATH = "apm_data.db"
FOLDER = os.path.join("..", "Historical prod data 25-26")
START_DATE = date(2025, 9, 30)
PLANT_TZ = "Europe/Athens"
ENTSOE_ZONE = "10YGR-HTSO-----Y"

_src = open("apm_app.py", encoding="utf-8").read()
_tree = ast.parse(_src)
def _extract(name):
    return next(ast.get_source_segment(_src, n) for n in _tree.body
               if isinstance(n, ast.FunctionDef) and n.name == name)

_ns = {
    "pd": pd, "sqlite3": sqlite3, "datetime": datetime, "timezone": timezone,
    "timedelta": timedelta, "date": date, "DB_PATH": DB_PATH,
    "PLANT_TZ": PLANT_TZ, "ENTSOE_ZONE": ENTSOE_ZONE, "Optional": Optional,
}
for _name in ("_get_db", "_read_cached_dam_prices", "_norm", "_compute_day_revenue_from_frames"):
    exec(compile(_extract(_name), f"<{_name}>", "exec"), _ns)

_get_db = _ns["_get_db"]
_read_cached_dam_prices = _ns["_read_cached_dam_prices"]
_compute_day_revenue_from_frames = _ns["_compute_day_revenue_from_frames"]


def _persist_day_revenue(d: date, r: dict) -> None:
    fetched_ts = datetime.now(timezone.utc).isoformat()
    conn = _get_db()
    conn.execute("""INSERT OR REPLACE INTO daily_revenue
        (day,kwh,revenue_eur,avg_price,fetched_ts) VALUES (?,?,?,?,?)""",
        (str(d), r["kwh"], r["revenue_eur"], r["avg_price"], fetched_ts))
    buckets = r.get("buckets")
    if buckets is not None and not buckets.empty:
        conn.executemany("""INSERT OR REPLACE INTO revenue_15min
            (dt,day,kwh,price_eur_mwh,revenue_eur,fetched_ts) VALUES (?,?,?,?,?,?)""",
            [(b["dt"].isoformat(), str(d), float(b["kwh"]), float(b["price"]),
             float(b["rev"]), fetched_ts) for _, b in buckets.iterrows()])
    conn.commit(); conn.close()


def load_local_production(folder: str) -> pd.DataFrame:
    """Returns DataFrame[dt (Athens tz-aware), kw_sum] — site-level power
    (sum of com1-1/com1-2/com1-3) at every 15-min timestamp present in the
    exports. Missing/no-sunlight timestamps are simply absent, not zero-
    filled — per instructions, not worth deep-diving."""
    ts_list, dst_list, kw_list = [], [], []
    files = sorted(glob.glob(os.path.join(folder, "Inverter_*.xlsx")))
    if not files:
        raise SystemExit(f"No Inverter_*.xlsx files found under {folder!r}")
    for fp in files:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=5, max_row=200000, max_col=8, values_only=True):
            start_time, power = row[3], row[4]
            if not start_time or power is None:
                continue
            is_dst = start_time.endswith(" DST")
            ts_list.append(start_time[:-4] if is_dst else start_time)
            dst_list.append(is_dst)
            kw_list.append(float(power))
        wb.close()

    df = pd.DataFrame({"ts_str": ts_list, "is_dst": dst_list, "kw": kw_list})
    naive = pd.to_datetime(df["ts_str"], format="%Y-%m-%d %H:%M:%S")
    df["dt"] = naive.dt.tz_localize(PLANT_TZ, ambiguous=df["is_dst"].to_numpy(),
                                    nonexistent="shift_forward")
    site = (df.groupby("dt")["kw"].sum().reset_index()
           .rename(columns={"kw": "kw_sum"}).sort_values("dt").reset_index(drop=True))
    return site


def _day_prod_json(site: pd.DataFrame, d: date) -> Optional[dict]:
    day_df = site[site["dt"].dt.date == d]
    if day_df.empty:
        return None
    rows = [{"collectTime": int(dt.tz_convert("UTC").timestamp() * 1000),
            "active_power": float(kw)} for dt, kw in zip(day_df["dt"], day_df["kw_sum"])]
    return {"data": rows, "_src": "15min"}


def compute_day(site: pd.DataFrame, d: date) -> Optional[dict]:
    pj = _day_prod_json(site, d)
    if pj is None:
        return None
    df_dam = _read_cached_dam_prices(d)
    if df_dam.empty:
        return None
    return _compute_day_revenue_from_frames(pj, df_dam[["dt", "price"]])


def get_existing_days() -> set:
    conn = sqlite3.connect(DB_PATH)
    days = {row[0] for row in conn.execute("SELECT day FROM daily_revenue").fetchall()}
    conn.close()
    return days


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--compare", help="YYYY-MM-DD: compute from Excel and compare vs cached daily_revenue, no write")
    ap.add_argument("--run", action="store_true", help="Backfill every day missing from daily_revenue")
    ap.add_argument("--fix-hourly", action="store_true",
                    help="Recompute days whose stored revenue_15min still looks hourly-fallback "
                         "(<=20 buckets/day) using local Excel data, wherever a file covers that day")
    args = ap.parse_args()

    print("Loading local production files...", flush=True)
    site = load_local_production(FOLDER)
    print(f"Loaded {len(site)} site-level 15-min readings, "
         f"{site['dt'].dt.date.nunique()} distinct days, "
         f"range {site['dt'].min()} .. {site['dt'].max()}", flush=True)

    if args.compare:
        d = date.fromisoformat(args.compare)
        r = compute_day(site, d)
        conn = sqlite3.connect(DB_PATH)
        cached = conn.execute("SELECT kwh, revenue_eur, avg_price FROM daily_revenue WHERE day=?",
                              (str(d),)).fetchone()
        conn.close()
        print(f"\nExcel-derived for {d}: {r['kwh']:.2f} kWh, EUR {r['revenue_eur']:.2f}, "
             f"avg_price {r['avg_price']:.2f}" if r else f"\nExcel-derived for {d}: no result")
        print(f"Cached (API)  for {d}: {cached}")
        if r and cached:
            diff = r["kwh"] - cached[0]
            pct = diff / cached[0] * 100 if cached[0] else float("nan")
            print(f"kWh diff: {diff:+.2f} ({pct:+.3f}%)")
        return

    if args.run:
        existing = get_existing_days()
        end_exclusive = date.today() + timedelta(days=1)
        d = START_DATE
        todo = []
        while d < end_exclusive:
            if str(d) not in existing:
                todo.append(d)
            d += timedelta(days=1)
        print(f"{len(todo)} day(s) missing from daily_revenue, attempting from local Excel", flush=True)

        ok, no_data, no_price = 0, [], []
        for i, day in enumerate(todo):
            r = compute_day(site, day)
            if r is None:
                df_dam = _read_cached_dam_prices(day)
                pj = _day_prod_json(site, day)
                if pj is None:
                    no_data.append(day)
                elif df_dam.empty:
                    no_price.append(day)
                else:
                    no_data.append(day)
                print(f"[{i+1}/{len(todo)}] {day}  SKIP (no usable data)", flush=True)
                continue
            _persist_day_revenue(day, r)
            ok += 1
            print(f"[{i+1}/{len(todo)}] {day}  OK  {r['kwh']:.1f} kWh, EUR {r['revenue_eur']:.2f}", flush=True)

        print(f"\nDone. {ok}/{len(todo)} days computed and persisted.")
        if no_data:
            print(f"{len(no_data)} day(s) with no local production data: {no_data}")
        if no_price:
            print(f"{len(no_price)} day(s) with no cached price data: {no_price}")
        return

    if args.fix_hourly:
        conn = sqlite3.connect(DB_PATH)
        rows = conn.execute("SELECT day, COUNT(*) c FROM revenue_15min GROUP BY day HAVING c<=20").fetchall()
        conn.close()
        candidates = [date.fromisoformat(r[0]) for r in rows]
        print(f"{len(candidates)} day(s) look hourly-fallback-derived (<=20 buckets)", flush=True)

        fixed, no_excel = 0, []
        for i, day in enumerate(candidates):
            r = compute_day(site, day)
            if r is None:
                no_excel.append(day)
                print(f"[{i+1}/{len(candidates)}] {day}  SKIP (no local Excel data for this day)", flush=True)
                continue
            _persist_day_revenue(day, r)
            fixed += 1
            print(f"[{i+1}/{len(candidates)}] {day}  FIXED  {r['kwh']:.1f} kWh, EUR {r['revenue_eur']:.2f}", flush=True)

        print(f"\nDone. {fixed}/{len(candidates)} days upgraded to native-resolution Excel data.")
        if no_excel:
            print(f"{len(no_excel)} day(s) left as hourly-fallback (no Excel file covers them): {no_excel}")
        return

    print("Specify --compare YYYY-MM-DD, --run, or --fix-hourly")


if __name__ == "__main__":
    main()
